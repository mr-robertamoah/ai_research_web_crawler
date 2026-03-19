"""
AmaliTech Competitor AI Services — Analysis Script (Ollama edition)
====================================================================
Drop-in replacement for analyse.py using a local Ollama model instead
of the Anthropic API. Everything else — incremental state tracking,
scoring, Excel output — is identical.

HOW TO RUN:
  python analyse_ollama.py                        # process only NEW site folders
  python analyse_ollama.py --competitor andela    # single competitor
  python analyse_ollama.py --max-pages 10         # faster / lower resource use
  python analyse_ollama.py --dry-run              # show state, no model calls
  python analyse_ollama.py --rerun-all            # reprocess everything

ENVIRONMENT VARIABLES:
  OLLAMA_HOST           Ollama base URL (default: http://host.docker.internal:11434)
  OLLAMA_MODEL          Model to use    (default: qwen3:8b)
  RERUN_ALL             set to 1 to reprocess everything
  WEIGHTS               comma-separated weights for 7 dimensions
  MAX_PAGES_PER_SITE    max pages to feed per competitor (default: 40)
  OCR_MIN_LENGTH        min chars for OCR row to be included (default: 30)
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import time
import textwrap
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("analyse_ollama")

# ── PATHS ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent.resolve()
SITES_DIR  = SCRIPT_DIR / "sites"
OUTPUT_DIR = SCRIPT_DIR / "output"
STATE_FILE = OUTPUT_DIR / "processed_folders_ollama.json"
MASTER_CSV = OUTPUT_DIR / "all_competitors_priority_ollama.csv"

# ── OLLAMA ────────────────────────────────────────────────────────────────────
OLLAMA_HOST  = os.getenv("OLLAMA_HOST",  "http://host.docker.internal:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "qwen3:8b")

# ── CATEGORIES / MATURITY / DIMENSIONS ───────────────────────────────────────
SERVICE_CATEGORIES = [
    "AI Advisory & Readiness",
    "AI Engineering & Automation",
    "AI Platforms & Agents",
    "AI-powered Solutions & New Revenue Models",
    "Talent & Staffing",
    "Other",
]

MATURITY_LEVELS = [
    "AI Explorer",
    "AI Practitioner",
    "AI Champion",
    "AI Explorer → AI Practitioner",
    "AI Practitioner → AI Champion",
]

DIMENSIONS = [
    "market_impact",
    "effort",
    "scalability",
    "revenue_potential",
    "market_credibility",
    "talent_availability",
    "strategic_fit",
]

DIM_LABELS = {
    "market_impact":       "Market Impact",
    "effort":              "Effort (inverse: 5=low effort)",
    "scalability":         "Scalability",
    "revenue_potential":   "Revenue Potential",
    "market_credibility":  "Market Credibility",
    "talent_availability": "Talent Availability",
    "strategic_fit":       "Strategic Fit",
}

DIM_DESCRIPTIONS = {
    "market_impact":       "Expected business value and demand. Strong client demand? Addresses major AI adoption challenge?",
    "effort":              "Difficulty to launch. SCORE 5 = VERY LOW effort (quick win). SCORE 1 = VERY HIGH effort.",
    "scalability":         "Can this be delivered repeatedly across multiple clients with standardised frameworks?",
    "revenue_potential":   "Typical deal size, follow-on work, recurring managed services opportunity.",
    "market_credibility":  "Would clients trust an offshore African provider (AmaliTech) to deliver this?",
    "talent_availability": "Skills available in Ghana/Rwanda, ease of hiring/training, cost of expertise.",
    "strategic_fit":       "Alignment with existing accounts (Schaeffler, Telekom, Knauf), capabilities, European compliance.",
}

_DIM_GUIDE = "\n".join(
    f"- {DIM_LABELS[d]}: {DIM_DESCRIPTIONS[d]}" for d in DIMENSIONS
)

# ── EXCEL COLOURS ─────────────────────────────────────────────────────────────
NAVY     = "1F3864"; MID_BLUE = "2E5DA8"; LIGHTER = "EBF3FA"; WHITE = "FFFFFF"
HIGH_BG  = "C6EFCE"; HIGH_FG = "276221"
MED_BG   = "FFEB9C"; MED_FG  = "9C6500"
LOW_BG   = "FFC7CE"; LOW_FG  = "9C0006"
GRAY     = "BFBFBF"

_thin  = Side(style="thin", color=GRAY)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ═════════════════════════════════════════════════════════════════════════════
# STATE
# ═════════════════════════════════════════════════════════════════════════════

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            log.warning("State file unreadable — starting fresh.")
    return {}

def save_state(state: dict) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")

def clear_state() -> None:
    if STATE_FILE.exists():
        STATE_FILE.unlink()
        log.info("State cleared — all folders will be reprocessed.")


# ═════════════════════════════════════════════════════════════════════════════
# WEIGHTS & SCORING
# ═════════════════════════════════════════════════════════════════════════════

def parse_weights(raw: str = "") -> dict[str, float]:
    raw = raw.strip() or os.getenv("WEIGHTS", "").strip()
    if raw:
        try:
            vals = [float(x.strip()) for x in raw.split(",")]
            if len(vals) == len(DIMENSIONS):
                return dict(zip(DIMENSIONS, vals))
        except (ValueError, TypeError):
            pass
    equal = 1.0 / len(DIMENSIONS)
    return {d: equal for d in DIMENSIONS}

def compute_score(scores: dict, weights: dict) -> float:
    total_w = sum(weights.values()) or 1
    weighted = sum(
        (scores.get(d, {}).get("score", 0) if isinstance(scores.get(d), dict) else 0)
        * weights[d]
        for d in DIMENSIONS
    )
    return round((weighted / (5 * total_w)) * 100, 1)

def priority_tier(score: float) -> str:
    if score >= 70: return "High"
    if score >= 45: return "Medium"
    return "Low"

def priority_display(score: float) -> str:
    return f"{priority_tier(score)} ({score})"


# ═════════════════════════════════════════════════════════════════════════════
# SITE DISCOVERY
# ═════════════════════════════════════════════════════════════════════════════

def find_site_folders(competitor_filter: str = "") -> list[Path]:
    if not SITES_DIR.exists():
        raise FileNotFoundError(f"sites/ not found at {SITES_DIR}.")
    folders = []
    for p in sorted(SITES_DIR.iterdir()):
        if not p.is_dir():
            continue
        has_content = (p / "pages_text.csv").exists() or (p / "posts_text.csv").exists()
        if not has_content:
            continue
        if competitor_filter and competitor_filter.lower() not in p.name.lower():
            continue
        folders.append(p)
    if not folders:
        raise FileNotFoundError(f"No site folders with content found under {SITES_DIR}.")
    return folders

def competitor_name(folder: Path) -> str:
    raw = folder.name.split("_")[0]
    raw = re.sub(r"-(com|ai|world|io|net|org|co)$", "", raw, flags=re.IGNORECASE)
    return raw.replace("-", " ").title()


# ═════════════════════════════════════════════════════════════════════════════
# CONTENT LOADING
# ═════════════════════════════════════════════════════════════════════════════

def load_content(folder: Path, max_pages: int, ocr_min_len: int) -> str:
    chunks: list[str] = []
    for csv_name in ("pages_text.csv", "posts_text.csv"):
        csv_path = folder / csv_name
        if not csv_path.exists():
            continue
        try:
            df = pd.read_csv(csv_path, dtype=str).fillna("").head(max_pages)
            for _, row in df.iterrows():
                url   = row.get("url", row.get("source_image", ""))
                title = row.get("page_title", row.get("source", ""))
                text  = row.get("clean_text", row.get("combined_text", row.get("provided_text", "")))
                if text.strip():
                    chunks.append(f"[PAGE: {title} | {url}]\n{text[:3000]}")
        except Exception as e:
            log.warning(f"Could not read {csv_name} in {folder.name}: {e}")

    ocr_path = folder / "ocr_output.csv"
    if ocr_path.exists():
        try:
            df = pd.read_csv(ocr_path, dtype=str).fillna("")
            for _, row in df.iterrows():
                ocr_text = row.get("extracted_text", "").strip()
                if len(ocr_text) >= ocr_min_len:
                    src = row.get("source_page_url", row.get("source_image", ""))
                    chunks.append(f"[IMAGE TEXT | {src}]\n{ocr_text}")
        except Exception as e:
            log.warning(f"Could not read ocr_output.csv in {folder.name}: {e}")

    return "\n\n---\n\n".join(chunks)

# ═════════════════════════════════════════════════════════════════════════════
# OLLAMA API
# ═════════════════════════════════════════════════════════════════════════════

def _call_ollama(system: str, user: str, retries: int = 3) -> str:
    """Call Ollama /api/chat endpoint and return the assistant message text."""
    url     = f"{OLLAMA_HOST}/api/chat"
    payload = {
        "model": OLLAMA_MODEL,
        "stream": False,
        "options": {"temperature": 0.1, "num_ctx": 8192},
        "messages": [
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
    }
    for attempt in range(retries):
        try:
            resp = requests.post(url, json=payload, timeout=600)
            resp.raise_for_status()
            return resp.json()["message"]["content"].strip()
        except requests.exceptions.ConnectionError:
            log.error(
                f"Cannot reach Ollama at {OLLAMA_HOST}. "
                "Make sure Ollama is running and OLLAMA_HOST is correct."
            )
            raise
        except Exception as e:
            log.warning(f"Ollama call failed (attempt {attempt+1}/{retries}): {e}")
            if attempt == retries - 1:
                raise
            time.sleep(5 * (attempt + 1))
    raise RuntimeError("All Ollama retries exhausted.")


def _parse_json(raw: str, context: str = "") -> dict | list | None:
    # Strip markdown code fences if present
    clean = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
    clean = re.sub(r"\s*```$", "", clean.strip(), flags=re.MULTILINE)
    # Also strip <think>...</think> blocks that qwen3 sometimes emits
    clean = re.sub(r"<think>.*?</think>", "", clean, flags=re.DOTALL).strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        # Try to find the first JSON array or object in the response
        for pattern in (r"(\[.*\])", r"(\{.*\})"):
            m = re.search(pattern, clean, re.DOTALL)
            if m:
                try:
                    return json.loads(m.group(1))
                except json.JSONDecodeError:
                    pass
        log.error(f"JSON parse error{' for ' + context if context else ''}. Snippet: {clean[:200]}")
        return None


# ── PROMPTS ───────────────────────────────────────────────────────────────────

_EXTRACTION_SYSTEM = textwrap.dedent(f"""
    You are a competitive intelligence analyst for AmaliTech — an AI-first
    technology services company from Ghana and Rwanda serving European enterprise
    clients in manufacturing, telecoms, and finance.

    Extract ALL AI-related and AI-adjacent services from the competitor content.
    Include: AI-branded services, ML engineering, data platforms, cloud AI,
    automation, analytics, RPA, AI talent placement, internal AI tools.

    Customer maturity levels:
    - AI Explorer: client just starting, no AI in place
    - AI Practitioner: client has some AI, actively scaling
    - AI Champion: AI-mature, needs advanced services
    - AI Explorer → AI Practitioner (transition)
    - AI Practitioner → AI Champion (transition)

    Return ONLY a valid JSON array. No markdown, no explanation, no preamble.
    Each element must have exactly these keys:
    {{
      "name": "short service name",
      "category": "one of: {' | '.join(SERVICE_CATEGORIES)}",
      "customer_maturity": "one of: {' | '.join(MATURITY_LEVELS)}",
      "description": "1-3 sentences on what this is and who it is for",
      "ai_classification": "core_ai or ai_adjacent",
      "evidence": "short quote or detail confirming this service",
      "source_url": "page URL where found, or empty string"
    }}

    If nothing relevant is found return: []
""").strip()

_SCORING_SYSTEM = textwrap.dedent(f"""
    You are a strategic analyst for AmaliTech scoring competitor services.

    AmaliTech context:
    - Delivery from Ghana + Rwanda; offshore/nearshore engineering + product management
    - Capabilities: Python/ML, Azure/cloud, Power BI, data engineering, QA, DevOps
    - Key accounts: Schaeffler (manufacturing), Deutsche Telekom (telco), Knauf, Serva
    - Target: United Internet (1&1) — telecoms and infrastructure AI
    - Differentiators: European compliance, data sovereignty, vendor-neutral AI, cost advantage

    Score each dimension 1–5. Ground every justification in AmaliTech's context.

    Dimensions:
    {_DIM_GUIDE}

    Return ONLY valid JSON, no markdown, no explanation:
    {{
      "market_impact":        {{"score": <1-5>, "justification": "<one sentence>"}},
      "effort":               {{"score": <1-5>, "justification": "<one sentence>"}},
      "scalability":          {{"score": <1-5>, "justification": "<one sentence>"}},
      "revenue_potential":    {{"score": <1-5>, "justification": "<one sentence>"}},
      "market_credibility":   {{"score": <1-5>, "justification": "<one sentence>"}},
      "talent_availability":  {{"score": <1-5>, "justification": "<one sentence>"}},
      "strategic_fit":        {{"score": <1-5>, "justification": "<one sentence>"}}
    }}
""").strip()


def extract_services(name: str, content: str) -> list[dict]:
    log.info("  Extracting services...")
    raw = _call_ollama(
        _EXTRACTION_SYSTEM,
        f"Competitor: {name}\n\nContent:\n{content[:20_000]}",
    )
    result = _parse_json(raw, context=name)
    if not isinstance(result, list):
        log.warning(f"  Unexpected extraction response for {name}.")
        return []
    log.info(f"  {len(result)} service(s) found.")
    return result


def score_service(competitor: str, service: dict) -> dict:
    user = (
        f"Competitor: {competitor}\n"
        f"Service: {service.get('name', '')}\n"
        f"Category: {service.get('category', '')}\n"
        f"Customer Maturity: {service.get('customer_maturity', '')}\n"
        f"Description: {service.get('description', '')}\n"
        f"AI Classification: {service.get('ai_classification', '')}\n\n"
        "Score this service against AmaliTech's priority matrix."
    )
    raw    = _call_ollama(_SCORING_SYSTEM, user)
    result = _parse_json(raw, context=service.get("name", ""))
    return result if isinstance(result, dict) else {}

# ═════════════════════════════════════════════════════════════════════════════
# ROW BUILDER
# ═════════════════════════════════════════════════════════════════════════════

def build_rows(competitor: str, services: list[dict],
               all_scores: list[dict], weights: dict) -> list[dict]:
    rows = []
    for svc, scores in zip(services, all_scores):
        score = compute_score(scores, weights)
        row: dict = {
            "competitor":        competitor,
            "service_name":      svc.get("name", ""),
            "category":          svc.get("category", ""),
            "customer_maturity": svc.get("customer_maturity", ""),
            "ai_classification": svc.get("ai_classification", ""),
            "description":       svc.get("description", ""),
            "evidence":          svc.get("evidence", ""),
            "source_url":        svc.get("source_url", ""),
            "priority_score":    score,
            "priority_tier":     priority_tier(score),
            "priority_display":  priority_display(score),
        }
        for dim in DIMENSIONS:
            d = scores.get(dim, {})
            row[f"{dim}_score"]         = d.get("score", "")         if isinstance(d, dict) else ""
            row[f"{dim}_justification"] = d.get("justification", "") if isinstance(d, dict) else ""
        rows.append(row)
    rows.sort(key=lambda r: r["priority_score"], reverse=True)
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# MASTER CSV
# ═════════════════════════════════════════════════════════════════════════════

def load_existing_results() -> pd.DataFrame:
    if MASTER_CSV.exists():
        try:
            df = pd.read_csv(MASTER_CSV, dtype=str).fillna("")
            log.info(f"Loaded {len(df)} existing rows from master CSV.")
            return df
        except Exception as e:
            log.warning(f"Could not read existing master CSV: {e}")
    return pd.DataFrame()

def merge_results(existing: pd.DataFrame, new_rows: list[dict]) -> pd.DataFrame:
    new_df = pd.DataFrame(new_rows) if new_rows else pd.DataFrame()
    if existing.empty:
        return new_df
    if new_df.empty:
        return existing
    new_competitors = new_df["competitor"].unique().tolist()
    existing_filtered = existing[~existing["competitor"].isin(new_competitors)]
    merged = pd.concat([existing_filtered, new_df], ignore_index=True)
    return merged.sort_values(["competitor", "priority_score"],
                              ascending=[True, False]).reset_index(drop=True)


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═════════════════════════════════════════════════════════════════════════════

def _write_long_list_sheet(ws, rows: list[dict], title: str) -> None:
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = title
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _fill(NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers    = ["#", "AI Service Category", "Customer Maturity", "Service", "Description", "Priority"]
    col_widths = [5, 32, 30, 34, 62, 18]
    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _align("center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for row_num, row in enumerate(rows, start=1):
        excel_row = row_num + 2
        tier = row.get("priority_tier", "")
        bg   = _fill(LIGHTER if row_num % 2 == 0 else WHITE)
        values = [row_num, row.get("category",""), row.get("customer_maturity",""),
                  row.get("service_name",""), row.get("description",""), row.get("priority_display","")]
        for col_i, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border = BORDER
            cell.font = _font(size=9)
            cell.alignment = _align("center" if col_i in (1,3,6) else "left", wrap=col_i==5)
            cell.fill = bg
        p = ws.cell(row=excel_row, column=6)
        if tier == "High":
            p.fill = _fill(HIGH_BG); p.font = _font(bold=True, color=HIGH_FG, size=9)
        elif tier == "Medium":
            p.fill = _fill(MED_BG);  p.font = _font(bold=True, color=MED_FG,  size=9)
        else:
            p.fill = _fill(LOW_BG);  p.font = _font(bold=True, color=LOW_FG,  size=9)
        ws.row_dimensions[excel_row].height = max(30, min(80, len(row.get("description","")) // 4))


def write_initiative_long_list(all_rows: list[dict], path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    by_competitor: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        by_competitor[row["competitor"]].append(row)

    ws_all = wb.create_sheet(title="All Competitors")
    combined = sorted(all_rows, key=lambda r: (
        r.get("competitor",""), r.get("category",""), -float(r.get("priority_score", 0))
    ))
    _write_long_list_sheet(ws_all, combined, "Initiative Long List — All Competitors")

    for comp_name in sorted(by_competitor.keys()):
        rows = sorted(by_competitor[comp_name],
                      key=lambda r: (r.get("category",""), -float(r.get("priority_score",0))))
        ws = wb.create_sheet(title=comp_name[:28])
        _write_long_list_sheet(ws, rows, f"Initiative Long List — {comp_name}")

    wb.save(path)
    log.info(f"  Initiative Long List: {path.name}  ({len(all_rows)} services, {len(by_competitor)} competitors)")


def write_detailed_workbook(competitor: str, rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scored Services"
    total_cols = 7 + len(DIMENSIONS) * 2
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    t = ws["A1"]
    t.value = f"AI Services Analysis — {competitor}"
    t.font = _font(bold=True, color=WHITE, size=13)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    base_headers = ["Service","Category","Customer Maturity","AI Classification",
                    "Description","Priority Score","Priority Tier"]
    dim_headers  = []
    for d in DIMENSIONS:
        dim_headers += [f"{DIM_LABELS[d]} Score", f"{DIM_LABELS[d]} Justification"]
    all_headers = base_headers + dim_headers
    col_widths  = [30,28,24,16,50,14,12] + [10,38]*len(DIMENSIONS)

    for col_i, (h, w) in enumerate(zip(all_headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font = _font(bold=True, color=WHITE, size=9)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _align("center", wrap=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    for row_num, row in enumerate(rows, start=1):
        excel_row = row_num + 2
        tier = row.get("priority_tier","")
        alt  = row_num % 2 == 0
        base_values = [row.get("service_name",""), row.get("category",""),
                       row.get("customer_maturity",""), row.get("ai_classification",""),
                       row.get("description",""), row.get("priority_score",""), row.get("priority_tier","")]
        dim_values = []
        for d in DIMENSIONS:
            dim_values += [row.get(f"{d}_score",""), row.get(f"{d}_justification","")]
        for col_i, value in enumerate(base_values + dim_values, start=1):
            cell = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border = BORDER
            cell.alignment = _align("left", wrap=True)
            cell.font = _font(size=9)
            cell.fill = _fill(LIGHTER if alt else WHITE)
        for col_i in (6, 7):
            cell = ws.cell(row=excel_row, column=col_i)
            if tier == "High":
                cell.fill = _fill(HIGH_BG); cell.font = _font(bold=True, color=HIGH_FG, size=9)
            elif tier == "Medium":
                cell.fill = _fill(MED_BG);  cell.font = _font(bold=True, color=MED_FG,  size=9)
            else:
                cell.fill = _fill(LOW_BG);  cell.font = _font(bold=True, color=LOW_FG,  size=9)
        ws.row_dimensions[excel_row].height = 40

    wb.save(path)
    log.info(f"  Detailed workbook: {path.name}")

# ═════════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═════════════════════════════════════════════════════════════════════════════

def _write_outputs(master: pd.DataFrame, state: dict) -> None:
    master.to_csv(MASTER_CSV, index=False, encoding="utf-8")
    log.info(f"  Master CSV: {MASTER_CSV.name}  ({len(master)} rows)")
    ts      = datetime.now().strftime("%Y%m%d")
    ll_path = OUTPUT_DIR / f"{ts}_initiative_long_list_ollama.xlsx"
    write_initiative_long_list(master.to_dict("records"), ll_path)
    competitors_done = sum(1 for v in state.values() if not v.get("skipped", False))
    log.info(f"\n{'='*60}")
    log.info(f"  Competitors processed : {competitors_done} / {len(state)} sites tracked")
    log.info(f"  Total services        : {len(master)}")
    for tier in ["High", "Medium", "Low"]:
        n = (master["priority_tier"] == tier).sum()
        log.info(f"  {tier:<8} priority    : {n}")
    log.info(f"  Outputs in            : {OUTPUT_DIR}/")
    log.info(f"{'='*60}\n")


def run(
    competitor_filter: str = "",
    weights_str: str = "",
    max_pages: int = 0,
    ocr_min_len: int = 0,
    dry_run: bool = False,
    rerun_all: bool = False,
) -> pd.DataFrame:
    rerun_all   = rerun_all or os.getenv("RERUN_ALL", "").strip() in ("1", "true", "yes")
    max_pages   = max_pages   or int(os.getenv("MAX_PAGES_PER_SITE", 40))
    ocr_min_len = ocr_min_len or int(os.getenv("OCR_MIN_LENGTH", 30))
    weights     = parse_weights(weights_str)

    OUTPUT_DIR.mkdir(exist_ok=True)

    if rerun_all:
        clear_state()
    state = load_state()

    all_folders = find_site_folders(competitor_filter)
    pending     = [f for f in all_folders if f.name not in state]
    completed   = [f for f in all_folders if f.name in state]

    log.info(f"\nOllama model      : {OLLAMA_MODEL}")
    log.info(f"Ollama host       : {OLLAMA_HOST}")
    log.info(f"Site folders      : {len(all_folders)}")
    log.info(f"  Already done    : {len(completed)}")
    log.info(f"  Pending         : {len(pending)}")

    if completed:
        log.info("  Completed:")
        for f in completed:
            info = state[f.name]
            log.info(f"    ✓ {competitor_name(f):<25} ({info.get('services','?')} services)")

    if pending:
        log.info("  Pending:")
        for f in pending:
            log.info(f"    … {competitor_name(f)}")

    if dry_run:
        log.info("\nDRY RUN — no model calls made.")
        return pd.DataFrame()

    if not pending:
        log.info("\nNo new folders to process. Rebuilding outputs from existing results...")
        existing = load_existing_results()
        if not existing.empty:
            _write_outputs(existing, state)
        return existing

    new_rows: list[dict] = []

    for folder in pending:
        name = competitor_name(folder)
        log.info(f"\n{'─'*60}")
        log.info(f"  {name}  ({folder.name})")
        log.info(f"{'─'*60}")

        content = load_content(folder, max_pages, ocr_min_len)
        if not content.strip():
            log.warning("  No content — skipping.")
            state[folder.name] = {
                "processed_at": datetime.now().isoformat(timespec="seconds"),
                "services": 0, "competitor": name, "skipped": True,
            }
            save_state(state)
            continue

        log.info(f"  Content: ~{len(content):,} chars")
        services = extract_services(name, content)
        if not services:
            log.warning("  No services extracted — skipping.")
            state[folder.name] = {
                "processed_at": datetime.now().isoformat(timespec="seconds"),
                "services": 0, "competitor": name, "skipped": True,
            }
            save_state(state)
            continue

        scored: list[dict] = []
        for i, svc in enumerate(services, 1):
            log.info(f"  Scoring [{i}/{len(services)}]: {svc.get('name','?')}")
            scores = score_service(name, svc)
            scored.append(scores)
            time.sleep(0.3)

        rows = build_rows(name, services, scored, weights)
        new_rows.extend(rows)

        safe        = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
        detail_path = OUTPUT_DIR / f"{safe}_services_scored_ollama.xlsx"
        write_detailed_workbook(name, rows, detail_path)

        state[folder.name] = {
            "processed_at": datetime.now().isoformat(timespec="seconds"),
            "services": len(rows), "competitor": name, "skipped": False,
        }
        save_state(state)
        log.info(f"  ✓ {name} done — {len(rows)} services scored.")

    existing = load_existing_results()
    master   = merge_results(existing, new_rows)
    if master.empty:
        log.warning("No results to write.")
        return master

    _write_outputs(master, state)
    return master


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="AmaliTech Competitor AI Services — Ollama Analysis"
    )
    parser.add_argument("--competitor", "-c", default="",
                        help="Filter to a single competitor (e.g. 'andela')")
    parser.add_argument("--weights", "-w", default="",
                        help="Comma-separated weights for 7 dimensions")
    parser.add_argument("--max-pages", "-p", type=int, default=0,
                        help="Max pages per competitor (default: 40)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show state without making model calls")
    parser.add_argument("--rerun-all", action="store_true",
                        help="Ignore state and reprocess everything")
    args = parser.parse_args()

    run(
        competitor_filter=args.competitor,
        weights_str=args.weights,
        max_pages=args.max_pages,
        dry_run=args.dry_run,
        rerun_all=args.rerun_all,
    )


if __name__ == "__main__":
    main()
