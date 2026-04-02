"""
analyse.py — unified analysis runner.

ANALYSE_MODE controls which pipeline to run:
  ANALYSE_MODE=competitor    → lib/pipelines/competitor.py  (from analyse_groq.py)
  ANALYSE_MODE=legacy        → lib/pipelines/legacy.py      (from legacy_analyse.py)
  ANALYSE_MODE=ai_consulting → lib/pipelines/ai_consulting.py (from ai_analyse.py)

Output files are always suffixed with the AI backend name (groq/claude)
so runs with different backends never overwrite each other.

Environment variables:
  ANALYSE_MODE      competitor | legacy | ai_consulting (default: competitor)
  AI_BACKEND        groq | claude (default: groq)
  GROQ_TIER         free | paid (default: free; ignored when AI_BACKEND=claude)
  GROQ_API_KEY      required if AI_BACKEND=groq
  GROQ_MODEL        default: llama-3.1-8b-instant
  ANTHROPIC_API_KEY required if AI_BACKEND=claude
  CLAUDE_MODEL      default: claude-haiku-4-20250514
  APP_DIR           base path (set to /app in Docker)
  SITES_DIR / LEGACY_DIR / AI_SITES_DIR   input dirs
  OUTPUT_DIR / LEGACY_OUTPUT_DIR / AI_OUTPUT_DIR  output dirs
  RERUN_ALL         set to 1 to reprocess everything
  WEIGHTS           comma-separated weights for 7 dimensions (competitor/legacy only)

Usage:
  python analyse_new.py
  ANALYSE_MODE=legacy python analyse_new.py --rerun-all
  ANALYSE_MODE=ai_consulting python analyse_new.py --source kpmg
  docker compose exec -e ANALYSE_MODE=competitor -e GROQ_API_KEY=... scraper python3 /app/input/analyse_new.py
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from lib.core import (
    AI_BACKEND, TIER_MAX_PAGES, call_ai,
    find_site_folders, source_name, load_content,
    load_state, save_state, clear_state,
)
from lib.excel import (
    fill, font, align, BORDER, NAVY, MID_BLUE, LIGHTER, WHITE,
    HIGH_BG, HIGH_FG, MED_BG, MED_FG, LOW_BG, LOW_FG,
    title_row, header_row, priority_cell, write_row,
)
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("analyse")

# ── MODE CONFIG ───────────────────────────────────────────────────────────────
ANALYSE_MODE = os.getenv("ANALYSE_MODE", "competitor").lower().strip()
SCRIPT_DIR   = Path(os.getenv("APP_DIR", Path(__file__).parent.resolve()))

_MODES = {
    "competitor":    ("SITES_DIR",    "sites",       "OUTPUT_DIR",          "output"),
    "legacy":        ("LEGACY_DIR",   "legacy",      "LEGACY_OUTPUT_DIR",   "legacy_output"),
    "ai_consulting": ("AI_SITES_DIR", "ai_sites",    "AI_OUTPUT_DIR",       "ai_output"),
    "client_intel":  ("CLIENT_SITES_DIR", "client_sites", "CLIENT_OUTPUT_DIR", "client_output"),
}
if ANALYSE_MODE not in _MODES:
    raise ValueError(f"Unknown ANALYSE_MODE '{ANALYSE_MODE}'. Choose: {list(_MODES)}")

_sites_env, _sites_default, _out_env, _out_default = _MODES[ANALYSE_MODE]
SITES_DIR  = Path(os.getenv(_sites_env,  str(SCRIPT_DIR / _sites_default)))
OUTPUT_DIR = Path(os.getenv(_out_env,    str(SCRIPT_DIR / _out_default)))
STATE_FILE = OUTPUT_DIR / f"{ANALYSE_MODE}_processed_{AI_BACKEND}.json"
MASTER_CSV = OUTPUT_DIR / f"{ANALYSE_MODE}_all_priority_{AI_BACKEND}.csv"

# Import the right pipeline
import importlib
_pipeline = importlib.import_module(f"lib.pipelines.{ANALYSE_MODE.replace('_consulting','_consulting').replace('competitor','competitor')}")
# Normalise module name
_pipeline = importlib.import_module({
    "competitor":    "lib.pipelines.competitor",
    "legacy":        "lib.pipelines.legacy",
    "ai_consulting": "lib.pipelines.ai_consulting",
    "client_intel":  "lib.pipelines.client_intel",
}[ANALYSE_MODE])


# ── WEIGHTS (competitor + legacy only) ────────────────────────────────────────
def parse_weights(raw: str = "") -> dict:
    dims = getattr(_pipeline, "DIMENSIONS", [])
    if not dims:
        return {}
    raw = raw.strip() or os.getenv("WEIGHTS", "").strip()
    if raw:
        try:
            vals = [float(x.strip()) for x in raw.split(",")]
            if len(vals) == len(dims):
                return dict(zip(dims, vals))
        except (ValueError, TypeError):
            pass
    equal = 1.0 / len(dims)
    return {d: equal for d in dims}


# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
def _write_sheet(ws, rows: list[dict], sheet_title: str) -> None:
    """Generic long-list sheet — columns adapt to mode."""
    if ANALYSE_MODE == "ai_consulting":
        headers = ["#","Source","Service Type","Service Name","Description",
                   "Delivery Format","Duration","Target Audience","Clients","Industries",
                   "Explicit Pricing","Inferred Price Range","Pricing Model","Pricing Confidence","Priority"]
        widths  = [5,20,20,28,45,22,14,20,28,28,18,20,16,16,14]
        def row_values(i, r):
            return [i, r.get("source",""), r.get("service_type",""), r.get("service_name",""),
                    r.get("description",""), r.get("delivery_format",""), r.get("duration",""),
                    r.get("target_audience",""), r.get("clients",""), r.get("industries",""),
                    r.get("pricing_explicit",""), r.get("inferred_price_range",""),
                    r.get("pricing_model",""), r.get("pricing_confidence",""), r.get("priority_tier","")]
        priority_col = 15; center_cols = {1, 15}
    elif ANALYSE_MODE == "legacy":
        headers = ["#","Source","Service/Product","Type","Category","Maturity","Description","Priority"]
        widths  = [5,22,30,12,26,14,45,18]
        def row_values(i, r):
            return [i, r.get("source",""), r.get("service_name",""), r.get("type",""),
                    r.get("category",""), r.get("maturity_level",""), r.get("description",""),
                    f"{r.get('priority_tier','')} ({r.get('priority_score','')})"]
        priority_col = 8; center_cols = {1, 4, 6, 8}
    else:  # competitor
        headers = ["#","Competitor","Category","Customer Maturity","Service","Description",
                   "Plain English","Clients","Industries","Priority"]
        widths  = [5,22,28,24,30,45,45,30,28,18]
        def row_values(i, r):
            return [i, r.get("competitor",""), r.get("category",""), r.get("customer_maturity",""),
                    r.get("service_name",""), r.get("description",""),
                    r.get("plain_english_summary",""), r.get("clients",""),
                    r.get("industries",""), r.get("priority_display","")]
        priority_col = 10; center_cols = {1, 3, 4, 10}

    if ANALYSE_MODE == "client_intel":
        headers = ["#","Client","Signal Type","Title","Description","Vendor/Tools",
                   "Budget Mention","Strategic Intent","Maturity","Source Type","Priority"]
        widths  = [5,22,18,30,45,30,22,35,14,18,12]
        def row_values(i, r):
            return [i, r.get("source",""), r.get("signal_type",""), r.get("title",""),
                    r.get("description",""), r.get("vendor_tools",""), r.get("budget_mention",""),
                    r.get("strategic_intent",""), r.get("maturity",""),
                    r.get("source_type",""), r.get("priority_tier","")]
        priority_col = 11; center_cols = {1, 3, 9, 10, 11}

    title_row(ws, 1, sheet_title, len(headers))
    header_row(ws, 2, headers, widths)
    ws.freeze_panes = "A3"
    for i, row in enumerate(rows, start=1):
        tier = row.get("priority_tier","")
        write_row(ws, i + 2, row_values(i, row), i,
                  center_cols=center_cols, priority_col=priority_col,
                  priority_tier=tier, row_height=40)


def _write_comparison_matrix_sheet(ws, all_rows: list[dict]) -> None:
    dims = getattr(_pipeline, "DIMENSIONS", [])
    from lib.pipelines.competitor import DIM_LABELS, priority_tier as pt
    headers = ["Competitor","Avg Score"] + [DIM_LABELS.get(d,d) for d in dims] + ["Confidence","# Services"]
    widths  = [28,14] + [14]*len(dims) + [14,12]
    title_row(ws, 1, "Competitor Comparison Matrix", len(headers))
    header_row(ws, 2, headers, widths)
    ws.freeze_panes = "A3"
    by_comp: dict[str, list] = defaultdict(list)
    for row in all_rows:
        by_comp[row.get("competitor","")].append(row)
    for i, comp in enumerate(sorted(by_comp), start=1):
        rows = by_comp[comp]
        avg  = round(sum(float(r.get("priority_score",0) or 0) for r in rows) / len(rows), 1)
        dim_avgs = []
        for d in dims:
            vals = [float(r[f"{d}_score"]) for r in rows if r.get(f"{d}_score") not in ("","",None)]
            dim_avgs.append(round(sum(vals)/len(vals),1) if vals else "")
        values = [comp, avg] + dim_avgs + [rows[0].get("data_confidence",""), len(rows)]
        write_row(ws, i+2, values, i, priority_col=2, priority_tier=pt(avg), row_height=20)


def _write_hypothesis_sheet(ws, hypothesis_data: dict) -> None:
    from lib.pipelines.competitor import HYPOTHESES
    headers = ["#","Hypothesis","Competitor","Evidence For","Evidence Against","Verdict"]
    widths  = [5,60,22,50,50,20]
    title_row(ws, 1, "Hypothesis Tracker", len(headers))
    header_row(ws, 2, headers, widths)
    ws.freeze_panes = "A3"
    er = 3
    for h_idx, hyp_text in enumerate(HYPOTHESES, start=1):
        h_key = f"h{h_idx}"
        entries = [(comp, data[h_key]) for comp, data in hypothesis_data.items() if h_key in data]
        if not entries:
            entries = [("—", {"evidence_for":"","evidence_against":"","verdict":"Insufficient data"})]
        for comp, entry in entries:
            verdict = entry.get("verdict","Insufficient data")
            bg_map  = {"Confirmed": HIGH_BG, "Refuted": LOW_BG}
            values  = [h_idx, hyp_text, comp, entry.get("evidence_for",""),
                       entry.get("evidence_against",""), verdict]
            write_row(ws, er, values, er, row_height=40)
            cell = ws.cell(row=er, column=6)
            cell.fill = fill(bg_map.get(verdict, MED_BG))
            cell.font = font(bold=True, size=9)
            er += 1


def write_long_list_xlsx(all_rows: list[dict], path: Path, hyp_data: dict = None) -> None:
    wb = Workbook(); wb.remove(wb.active)
    ws_all = wb.create_sheet("All Sources")
    _write_sheet(ws_all, sorted(all_rows, key=lambda r: (
        r.get("source", r.get("competitor","")), -float(r.get("priority_score",0) or 0)
    )), f"{ANALYSE_MODE.replace('_',' ').title()} — All Sources")

    # Per-source/competitor sheets
    key = "competitor" if ANALYSE_MODE == "competitor" else "source"
    by_src: dict[str, list] = defaultdict(list)
    for row in all_rows:
        by_src[row.get(key,"")].append(row)
    for src in sorted(by_src):
        ws = wb.create_sheet(src[:28])
        _write_sheet(ws, sorted(by_src[src], key=lambda r: -float(r.get("priority_score",0) or 0)),
                     f"{src}")

    # Comparison sheet for ai_consulting
    if ANALYSE_MODE == "ai_consulting" and hasattr(_pipeline, "SERVICE_TYPES"):
        ws_cmp = wb.create_sheet("By Service Type")
        _write_comparison_sheet(ws_cmp, all_rows)

    # Competitor-specific sheets
    if ANALYSE_MODE == "competitor":
        ws_matrix = wb.create_sheet("Comparison Matrix")
        _write_comparison_matrix_sheet(ws_matrix, all_rows)
        if hyp_data:
            ws_hyp = wb.create_sheet("Hypothesis Tracker")
            _write_hypothesis_sheet(ws_hyp, hyp_data)

    wb.save(path)
    log.info(f"  Long list saved: {path.name}  ({len(all_rows)} items)")


def _write_comparison_sheet(ws, rows: list[dict]) -> None:
    title_row(ws, 1, "AI Services — Comparison by Service Type", 8)
    header_row(ws, 2, ["Service Type","Company","Service Name","Description","Format / Duration","Clients","Industries","Pricing"], [22,20,28,45,24,28,28,22])
    ws.freeze_panes = "A3"
    grouped: dict[str, list] = defaultdict(list)
    for row in rows:
        grouped[row.get("service_type","Other")].append(row)
    er = 3
    for stype in _pipeline.SERVICE_TYPES:
        group_rows = grouped.get(stype, [])
        if not group_rows:
            continue
        ws.merge_cells(f"A{er}:H{er}")
        c = ws.cell(row=er, column=1, value=stype)
        c.font = font(bold=True, color=WHITE, size=10); c.fill = fill(MID_BLUE); c.alignment = align("left")
        ws.row_dimensions[er].height = 20; er += 1
        for row in sorted(group_rows, key=lambda r: r.get("company","")):
            pricing = row.get("pricing_explicit","") or row.get("inferred_price_range","")
            if row.get("pricing_confidence","") not in ("explicit",""):
                pricing += " (est.)"
            values = [stype, row.get("company", row.get("source","")), row.get("service_name",""),
                      row.get("description",""),
                      f"{row.get('delivery_format','')} {row.get('duration','')}".strip(),
                      row.get("clients",""), row.get("industries",""), pricing]
            write_row(ws, er, values, er, row_height=40); er += 1


# ── MERGE ─────────────────────────────────────────────────────────────────────
def _key_col():
    return "competitor" if ANALYSE_MODE == "competitor" else "source"


def load_existing() -> pd.DataFrame:
    if MASTER_CSV.exists():
        try:
            df = pd.read_csv(MASTER_CSV, dtype=str).fillna("")
            log.info(f"Loaded {len(df)} existing rows from master CSV.")
            return df
        except Exception as e:
            log.warning(f"Could not read master CSV: {e}")
    return pd.DataFrame()


def merge_results(existing: pd.DataFrame, new_rows: list[dict]) -> pd.DataFrame:
    new_df = pd.DataFrame(new_rows) if new_rows else pd.DataFrame()
    if existing.empty: return new_df
    if new_df.empty:   return existing
    key = _key_col()
    new_keys = new_df[key].unique().tolist()
    filtered = existing[~existing[key].isin(new_keys)]
    merged   = pd.concat([filtered, new_df], ignore_index=True)
    return merged.sort_values([key, "priority_score"], ascending=[True, False]).reset_index(drop=True)


# ── MAIN PIPELINE ─────────────────────────────────────────────────────────────
def run(source_filter: str = "", weights_str: str = "", max_pages: int = 0,
        dry_run: bool = False, rerun_all: bool = False) -> pd.DataFrame:

    max_pages = max_pages or int(os.getenv("MAX_PAGES_PER_SITE", TIER_MAX_PAGES))
    rerun_all = rerun_all or os.getenv("RERUN_ALL","").strip() in ("1","true","yes")
    weights   = parse_weights(weights_str)
    OUTPUT_DIR.mkdir(exist_ok=True)

    if rerun_all:
        clear_state(STATE_FILE)
    state = load_state(STATE_FILE)

    all_folders = find_site_folders(SITES_DIR, source_filter)
    pending     = [f for f in all_folders if f.name not in state]
    completed   = [f for f in all_folders if f.name in state]

    log.info(f"\nMode        : {ANALYSE_MODE.upper()}")
    log.info(f"AI backend  : {AI_BACKEND.upper()}")
    log.info(f"Folders     : {len(all_folders)}  (done={len(completed)}, pending={len(pending)})")

    if dry_run:
        for f in pending:
            log.info(f"  … {source_name(f)}")
        log.info("DRY RUN — no API calls.")
        return pd.DataFrame()

    all_content_chunks: list[str] = []
    for folder in completed:
        chunk = load_content(folder, max_pages)
        if chunk:
            all_content_chunks.append(f"=== {source_name(folder)} ===\n{chunk}")

    new_rows: list[dict] = []
    hyp_data: dict = {}
    hyp_file = OUTPUT_DIR / f"hypothesis_tracker_{AI_BACKEND}.json"
    if ANALYSE_MODE == "competitor":
        hyp_data = json.loads(hyp_file.read_text()) if hyp_file.exists() else {}

    def save_hypothesis_data(data):
        hyp_file.write_text(json.dumps(data, indent=2, ensure_ascii=False))

    if not pending:
        log.info("Nothing new to process — rebuilding outputs.")
        existing = load_existing()
        if not existing.empty:
            _rebuild_outputs(existing, state, all_folders, max_pages, weights, hyp_data)
        return existing

    for folder in pending:
        name = source_name(folder)
        log.info(f"\n{'─'*55}\n  {name}  ({folder.name})\n{'─'*55}")

        content = load_content(folder, max_pages)
        if not content.strip():
            log.warning("  No content — skipping.")
            state[folder.name] = {"processed_at": datetime.now().isoformat(timespec="seconds"),
                                  "services": 0, "source": name, "skipped": True}
            save_state(STATE_FILE, state); continue

        log.info(f"  Content: ~{len(content):,} chars")
        all_content_chunks.append(f"=== {name} ===\n{content}")

        services = _pipeline.extract_services(name, content)
        if not services:
            log.warning("  No services extracted — skipping.")
            state[folder.name] = {"processed_at": datetime.now().isoformat(timespec="seconds"),
                                  "services": 0, "source": name, "skipped": True}
            save_state(STATE_FILE, state); continue

        # Score / build rows
        if ANALYSE_MODE == "competitor":
            conf = _pipeline.data_confidence(content)
            hyp  = _pipeline.assess_hypotheses(name, content)
            hyp_data[name] = hyp
            save_hypothesis_data(hyp_data)
            all_scores = []
            for i, svc in enumerate(services, 1):
                log.info(f"  Scoring [{i}/{len(services)}]: {svc.get('name','?')}")
                all_scores.append(_pipeline.score_service(name, svc))
                time.sleep(0.3)
            rows = _pipeline.build_rows(name, services, all_scores, weights, confidence=conf)
        elif ANALYSE_MODE == "legacy":
            all_scores = []
            for i, svc in enumerate(services, 1):
                log.info(f"  Scoring [{i}/{len(services)}]: {svc.get('name','?')}")
                all_scores.append(_pipeline.score_service(name, svc))
                time.sleep(0.3)
            rows = _pipeline.build_rows(name, services, all_scores, weights)
        else:  # ai_consulting + client_intel
            rows = _pipeline.build_rows(name, services)

        new_rows.extend(rows)

        # Per-source workbook
        safe        = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
        detail_path = OUTPUT_DIR / f"{safe}_{ANALYSE_MODE}_{AI_BACKEND}.xlsx"
        _write_detail_workbook(name, rows, detail_path)

        state[folder.name] = {"processed_at": datetime.now().isoformat(timespec="seconds"),
                              "services": len(rows), "source": name, "skipped": False}
        save_state(STATE_FILE, state)

        # Incremental flush
        partial = merge_results(load_existing(), rows)
        partial.to_csv(MASTER_CSV, index=False, encoding="utf-8")
        log.info(f"  ✓ {name} — {len(rows)} services.")

    existing = load_existing()
    master   = merge_results(existing, new_rows)
    if master.empty:
        log.warning("No results to write.")
        return master

    log.info("  Waiting 30s before brief generation to clear rate limit...")
    time.sleep(30)
    _rebuild_outputs(master, state, all_folders, max_pages, weights, hyp_data,
                     all_content_chunks=all_content_chunks)
    return master


def _rebuild_outputs(master: pd.DataFrame, state: dict, all_folders: list,
                     max_pages: int, weights: dict, hyp_data: dict,
                     all_content_chunks: list = None) -> None:
    master.to_csv(MASTER_CSV, index=False, encoding="utf-8")
    ts       = datetime.now().strftime("%Y%m%d")
    all_rows = master.to_dict("records")

    write_long_list_xlsx(all_rows, OUTPUT_DIR / f"{ts}_{ANALYSE_MODE}_long_list_{AI_BACKEND}.xlsx",
                         hyp_data=hyp_data if ANALYSE_MODE == "competitor" else None)

    # Generate brief
    if not all_content_chunks:
        all_content_chunks = [
            f"=== {source_name(f)} ===\n{load_content(f, max_pages)}"
            for f in all_folders if load_content(f, max_pages).strip()
        ]
    combined = "\n\n".join(all_content_chunks)
    brief = {}
    if hasattr(_pipeline, "generate_brief"):
        brief = _pipeline.generate_brief(combined)
        if brief and hasattr(_pipeline, "write_brief_md"):
            _pipeline.write_brief_md(brief, all_rows, OUTPUT_DIR / f"{ts}_{ANALYSE_MODE}_brief_{AI_BACKEND}.md")

    if ANALYSE_MODE == "ai_consulting":
        if hasattr(_pipeline, "write_market_summary_md"):
            _pipeline.write_market_summary_md(all_rows, OUTPUT_DIR / f"{ts}_ai_market_summary_{AI_BACKEND}.md")
        if hasattr(_pipeline, "write_executive_brief_md"):
            _pipeline.write_executive_brief_md(brief, all_rows, OUTPUT_DIR / f"{ts}_ai_executive_brief_{AI_BACKEND}.md")

    if ANALYSE_MODE == "competitor":
        if hasattr(_pipeline, "write_market_summary_md"):
            _pipeline.write_market_summary_md(all_rows, OUTPUT_DIR / f"{ts}_competitor_market_summary_{AI_BACKEND}.md")
        if hasattr(_pipeline, "write_executive_brief_md"):
            _pipeline.write_executive_brief_md(brief, all_rows, OUTPUT_DIR / f"{ts}_competitor_executive_brief_{AI_BACKEND}.md")

    if ANALYSE_MODE == "client_intel":
        if hasattr(_pipeline, "write_market_summary_md"):
            _pipeline.write_market_summary_md(all_rows, OUTPUT_DIR / f"{ts}_client_market_summary_{AI_BACKEND}.md")
        if hasattr(_pipeline, "write_executive_brief_md"):
            _pipeline.write_executive_brief_md(brief, all_rows, OUTPUT_DIR / f"{ts}_client_executive_brief_{AI_BACKEND}.md")
        if hasattr(_pipeline, "write_potential_clients_md"):
            _pipeline.write_potential_clients_md(all_rows, all_folders, OUTPUT_DIR / f"{ts}_potential_clients_{AI_BACKEND}.md")

    done = sum(1 for v in state.values() if not v.get("skipped", False))
    log.info(f"\n{'='*55}")
    log.info(f"  Mode              : {ANALYSE_MODE}")
    log.info(f"  Sources processed : {done} / {len(state)}")
    log.info(f"  Total services    : {len(master)}")
    for tier in ["High","Medium","Low"]:
        if "priority_tier" in master.columns:
            log.info(f"  {tier:<8} : {(master['priority_tier'] == tier).sum()}")
    log.info(f"  Outputs in        : {OUTPUT_DIR}/")
    log.info(f"{'='*55}\n")


def _write_detail_workbook(source: str, rows: list[dict], path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Scored Services"
    if ANALYSE_MODE == "ai_consulting":
        headers = ["Service Name","Company","Service Type","Description","Delivery Format",
                   "Duration","Target Audience","Explicit Pricing","Inferred Price Range",
                   "Pricing Model","Pricing Confidence","Pricing Basis","Evidence","Source URL","Priority"]
        widths  = [28,20,20,45,22,14,20,18,20,16,16,30,35,30,12]
        def vals(r): return [r.get("service_name",""), r.get("company",source), r.get("service_type",""),
                             r.get("description",""), r.get("delivery_format",""), r.get("duration",""),
                             r.get("target_audience",""), r.get("pricing_explicit",""),
                             r.get("inferred_price_range",""), r.get("pricing_model",""),
                             r.get("pricing_confidence",""), r.get("pricing_basis",""),
                             r.get("evidence",""), r.get("source_url",""), r.get("priority_tier","")]
        pcol = 15
    elif ANALYSE_MODE == "legacy":
        dims = getattr(_pipeline, "DIMENSIONS", [])
        base_h = ["Service","Source","Type","Category","Maturity","Description","Evidence","Priority Score","Priority Tier"]
        dim_h  = [f"{d} Score" for d in dims] + [f"{d} Justification" for d in dims]
        headers = base_h + dim_h
        widths  = [30,20,12,26,14,45,35,14,12] + [10]*len(dims) + [36]*len(dims)
        def vals(r):
            base = [r.get("service_name",""), r.get("source",""), r.get("type",""), r.get("category",""),
                    r.get("maturity_level",""), r.get("description",""), r.get("evidence",""),
                    r.get("priority_score",""), r.get("priority_tier","")]
            dim_scores = [r.get(f"{d}_score","") for d in dims]
            dim_justs  = [r.get(f"{d}_justification","") for d in dims]
            return base + dim_scores + dim_justs
        pcol = 9
    else:  # competitor
        dims = getattr(_pipeline, "DIMENSIONS", [])
        base_h = ["Service","Competitor","Category","Customer Maturity","AI Class","Description",
                  "Plain English","Pricing Signals","Clients","Industries","Tech Stack","Confidence",
                  "Evidence","Source URL","Priority Score","Priority Tier"]
        dim_h  = [f"{d} Score" for d in dims] + [f"{d} Justification" for d in dims]
        headers = base_h + dim_h
        widths  = [30,20,26,22,12,45,45,22,30,28,22,10,35,30,14,12] + [10]*len(dims) + [36]*len(dims)
        def vals(r):
            base = [r.get("service_name",""), r.get("competitor",""), r.get("category",""),
                    r.get("customer_maturity",""), r.get("ai_classification",""), r.get("description",""),
                    r.get("plain_english_summary",""), r.get("pricing_signals",""),
                    r.get("clients",""), r.get("industries",""),
                    r.get("tech_stack",""), r.get("data_confidence",""), r.get("evidence",""),
                    r.get("source_url",""), r.get("priority_score",""), r.get("priority_tier","")]
            dim_scores = [r.get(f"{d}_score","") for d in dims]
            dim_justs  = [r.get(f"{d}_justification","") for d in dims]
            return base + dim_scores + dim_justs
        pcol = 16

    if ANALYSE_MODE == "client_intel":
        headers = ["Client","Signal Type","Title","Description","Vendor/Tools","Budget Mention",
                   "Strategic Intent","Maturity","Source Type","Evidence","Source URL","Priority"]
        widths  = [22,18,30,45,30,22,35,14,18,40,30,12]
        def vals(r): return [r.get("source",""), r.get("signal_type",""), r.get("title",""),
                             r.get("description",""), r.get("vendor_tools",""), r.get("budget_mention",""),
                             r.get("strategic_intent",""), r.get("maturity",""), r.get("source_type",""),
                             r.get("evidence",""), r.get("source_url",""), r.get("priority_tier","")]
        pcol = 12

    title_row(ws, 1, f"{ANALYSE_MODE.replace('_',' ').title()} Analysis — {source}", len(headers))
    header_row(ws, 2, headers, widths)
    ws.freeze_panes = "A3"
    for i, row in enumerate(rows, start=1):
        write_row(ws, i + 2, vals(row), i, priority_col=pcol,
                  priority_tier=row.get("priority_tier",""), row_height=45)
    wb.save(path)
    log.info(f"  Detailed workbook: {path.name}")


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="AmaliTech Unified Analysis Runner")
    parser.add_argument("--source", "-s", default="", help="Filter to one source/competitor")
    parser.add_argument("--weights", "-w", default="", help="Comma-separated dimension weights")
    parser.add_argument("--max-pages", "-p", type=int, default=0)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--rerun-all", action="store_true")
    args = parser.parse_args()
    run(source_filter=args.source, weights_str=args.weights, max_pages=args.max_pages,
        dry_run=args.dry_run, rerun_all=args.rerun_all)


if __name__ == "__main__":
    main()
