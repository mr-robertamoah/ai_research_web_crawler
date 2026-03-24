"""
Legacy Modernization Analysis Script
======================================
Reads scraped output from legacy_scraper.py, extracts legacy modernisation
services and products, scores them against the priority matrix, answers the
pillar lead's four research questions, and produces:

  output/
    YYYYMMDD_legacy_long_list.xlsx       — Initiative Long List (competitor + service)
    YYYYMMDD_legacy_research_brief.md    — Research brief answering the 4 questions
    legacy_all_priority.csv              — Flat master CSV for further processing
    <source>_legacy_scored.xlsx          — Per-source detailed workbook

FOCUS AREAS (per pillar lead request):
  1. Who is doing AI-assisted legacy / mainframe / COBOL modernisation?
  2. Does it work — what is the maturity level?
  3. What state-of-the-art tools and approaches exist (including MIT/academic)?
  4. Java 8/11 → 17/21 migration using AI as a supportive hand.

AI BACKEND (controlled by env var):
  AI_BACKEND=groq     — uses Groq API (default)
  AI_BACKEND=claude   — uses Anthropic Claude API

HOW TO RUN:
  python legacy_analyse.py                        # all sites, Groq, equal weights
  AI_BACKEND=claude python legacy_analyse.py      # use Claude instead
  python legacy_analyse.py --source capgemini     # single source
  python legacy_analyse.py --dry-run              # show pending, no API calls
  python legacy_analyse.py --rerun-all            # reprocess everything
  python legacy_analyse.py --max-pages 15         # limit pages per source

ENVIRONMENT VARIABLES:
  AI_BACKEND            groq | claude (default: groq)
  GROQ_API_KEY          required if AI_BACKEND=groq
  GROQ_MODEL            Groq model (default: qwen/qwen3-32b)
  ANTHROPIC_API_KEY     required if AI_BACKEND=claude
  RERUN_ALL             set to 1 to reprocess everything
  WEIGHTS               comma-separated weights for 7 dimensions (default: equal)
  MAX_PAGES_PER_SITE    max pages to feed per source (default: 40)
  OCR_MIN_LENGTH        min chars for OCR row inclusion (default: 30)

WEIGHT ORDER:
  market_impact, effort, scalability, revenue_potential,
  market_credibility, talent_availability, strategic_fit

EFFORT IS INVERSE: score 5 = low effort (fast to launch), score 1 = very high.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import time
import textwrap
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("legacy_analyse")

# ── PATHS ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(os.getenv("APP_DIR", Path(__file__).parent.resolve()))
SITES_DIR   = Path(os.getenv("LEGACY_DIR", "/app/legacy"))
OUTPUT_DIR  = Path(os.getenv("LEGACY_OUTPUT_DIR", SCRIPT_DIR / "legacy_output"))
STATE_FILE  = OUTPUT_DIR / "legacy_processed_folders.json"
MASTER_CSV  = OUTPUT_DIR / "legacy_all_priority.csv"

# ── AI BACKEND CONFIG ─────────────────────────────────────────────────────────
AI_BACKEND   = os.getenv("AI_BACKEND", "groq").lower().strip()
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
GROQ_MODEL   = os.getenv("GROQ_MODEL", "qwen/qwen3-32b")
GROQ_URL     = "https://api.groq.com/openai/v1/chat/completions"
CLAUDE_KEY   = os.getenv("ANTHROPIC_API_KEY", "")
CLAUDE_MODEL = "claude-sonnet-4-20250514"

# ── SERVICE CATEGORIES ────────────────────────────────────────────────────────
SERVICE_CATEGORIES = [
    "AI Advisory & Readiness",
    "AI Engineering & Automation",
    "AI Platforms & Agents",
    "AI-powered Solutions & New Revenue Models",
    "Talent & Staffing",
    "Other",
]

MATURITY_LEVELS = [
    "AI Explorer",
    "AI Practitioner",
    "AI Champion",
    "AI Explorer → AI Practitioner",
    "AI Practitioner → AI Champion",
]

# ── PRIORITY MATRIX ───────────────────────────────────────────────────────────
DIMENSIONS = [
    "market_impact",
    "effort",
    "scalability",
    "revenue_potential",
    "market_credibility",
    "talent_availability",
    "strategic_fit",
]

DIM_LABELS = {
    "market_impact":       "Market Impact",
    "effort":              "Effort (inverse: 5=low effort)",
    "scalability":         "Scalability",
    "revenue_potential":   "Revenue Potential",
    "market_credibility":  "Market Credibility",
    "talent_availability": "Talent Availability",
    "strategic_fit":       "Strategic Fit",
}

DIM_DESCRIPTIONS = {
    "market_impact": (
        "Demand for legacy modernisation services. Are European enterprises actively "
        "looking to modernise COBOL/mainframe or upgrade Java versions? "
        "Does AI assistance make this a compelling offering?"
    ),
    "effort": (
        "Difficulty for AmaliTech to deliver this. Consider: specialised COBOL/mainframe "
        "expertise required, tooling needed, time to build capability. "
        "SCORE 5 = very low effort (quick win). SCORE 1 = very high effort."
    ),
    "scalability": (
        "Can AmaliTech deliver this repeatedly across multiple clients using "
        "standardised frameworks, AI tools (Amazon Q, OpenRewrite, watsonx), "
        "and reusable playbooks?"
    ),
    "revenue_potential": (
        "Deal size and recurring opportunity. Large mainframe estates = large projects. "
        "Java migration can be recurring as clients upgrade through LTS versions. "
        "Potential for managed modernisation retainers."
    ),
    "market_credibility": (
        "Would European enterprise clients trust an offshore African provider for "
        "mission-critical legacy modernisation? Consider TISAX/ISO 27001 certifications, "
        "existing manufacturing and telco account relationships."
    ),
    "talent_availability": (
        "Can AmaliTech staff this? Ghana/Rwanda talent pool for Java, Python, "
        "cloud-native development is strong. COBOL expertise is rare everywhere — "
        "but AI tools reduce the need for deep COBOL knowledge."
    ),
    "strategic_fit": (
        "Alignment with AmaliTech's position: existing Java/Python/Azure capability; "
        "manufacturing accounts (Schaeffler, Knauf) likely have Java estates; "
        "telco accounts (Telekom, 1&1) may have legacy systems; "
        "European compliance positioning; AI-first services strategy."
    ),
}

_DIM_GUIDE = "\n".join(
    f"- {DIM_LABELS[d]}: {DIM_DESCRIPTIONS[d]}" for d in DIMENSIONS
)

# ── RESEARCH QUESTIONS (pillar lead) ─────────────────────────────────────────
RESEARCH_QUESTIONS = [
    {
        "id":       "Q1",
        "question": "Who is doing AI-assisted legacy/mainframe/COBOL modernisation — "
                    "which companies offer this as a service or product?",
    },
    {
        "id":       "Q2",
        "question": "Does it work — what is the maturity level of AI-assisted legacy "
                    "modernisation? Are there proven results or is it still experimental?",
    },
    {
        "id":       "Q3",
        "question": "What state-of-the-art tools, platforms, and approaches exist — "
                    "including any academic or research backing (MIT, IBM Research, etc.)?",
    },
    {
        "id":       "Q4",
        "question": "What specifically exists for Java 8/11 to Java 17/21 migration "
                    "using AI as a supportive hand — tools, services, and proven results?",
    },
]

# ── EXCEL COLOURS ─────────────────────────────────────────────────────────────
NAVY     = "1F3864"; MID_BLUE = "2E5DA8"; LIGHTER = "EBF3FA"; WHITE = "FFFFFF"
HIGH_BG  = "C6EFCE"; HIGH_FG = "276221"
MED_BG   = "FFEB9C"; MED_FG  = "9C6500"
LOW_BG   = "FFC7CE"; LOW_FG  = "9C0006"
GRAY     = "BFBFBF"

_thin  = Side(style="thin", color=GRAY)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _fill(h):  return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ═════════════════════════════════════════════════════════════════════════════
# AI BACKEND — unified call interface
# ═════════════════════════════════════════════════════════════════════════════

def _call_groq(system: str, user: str, max_tokens: int = 4096,
               retries: int = 3) -> str:
    import requests as req
    if not GROQ_API_KEY:
        raise EnvironmentError(
            "GROQ_API_KEY is not set. Run: export GROQ_API_KEY=your_key"
        )
    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type":  "application/json",
    }
    payload = {
        "model":      GROQ_MODEL,
        "max_tokens": max_tokens,
        "messages": [
            {"role": "system",  "content": system},
            {"role": "user",    "content": user},
        ],
    }
    for attempt in range(retries):
        try:
            resp = req.post(GROQ_URL, headers=headers,
                            json=payload, timeout=60)
            resp.raise_for_status()
            return resp.json()["choices"][0]["message"]["content"].strip()
        except Exception as e:
            if attempt == retries - 1:
                raise
            wait = 20 * (attempt + 1)
            log.warning(f"Groq error ({e}) — retrying in {wait}s")
            time.sleep(wait)
    raise RuntimeError("Groq: all retries exhausted.")


def _call_claude(system: str, user: str, max_tokens: int = 4096,
                 retries: int = 3) -> str:
    try:
        import anthropic
    except ImportError:
        raise ImportError("Run: pip install anthropic")
    if not CLAUDE_KEY:
        raise EnvironmentError(
            "ANTHROPIC_API_KEY is not set. Run: export ANTHROPIC_API_KEY=your_key"
        )
    client = anthropic.Anthropic(api_key=CLAUDE_KEY)
    for attempt in range(retries):
        try:
            resp = client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=max_tokens,
                system=system,
                messages=[{"role": "user", "content": user}],
            )
            return resp.content[0].text.strip()
        except Exception as e:
            if "rate" in str(e).lower():
                wait = 30 * (attempt + 1)
                log.warning(f"Claude rate limit — waiting {wait}s")
                time.sleep(wait)
            elif attempt == retries - 1:
                raise
            else:
                time.sleep(10)
    raise RuntimeError("Claude: all retries exhausted.")


def call_ai(system: str, user: str, max_tokens: int = 4096) -> str:
    """Route to Groq or Claude based on AI_BACKEND env var."""
    if AI_BACKEND == "claude":
        return _call_claude(system, user, max_tokens)
    return _call_groq(system, user, max_tokens)


def _parse_json(raw: str, context: str = "") -> dict | list | None:
    clean = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
    # Strip Qwen <think>...</think> blocks if present
    clean = re.sub(r"<think>.*?</think>", "", clean, flags=re.DOTALL).strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError as e:
        log.error(f"JSON parse error{' [' + context + ']' if context else ''}: {e}")
        log.debug(f"Raw snippet: {clean[:400]}")
        return None


# ═════════════════════════════════════════════════════════════════════════════
# STATE
# ═════════════════════════════════════════════════════════════════════════════

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            log.warning("State file unreadable — starting fresh.")
    return {}


def save_state(state: dict) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    STATE_FILE.write_text(
        json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def clear_state() -> None:
    if STATE_FILE.exists():
        STATE_FILE.unlink()
        log.info("Legacy state cleared — all folders will be reprocessed.")


# ═════════════════════════════════════════════════════════════════════════════
# WEIGHTS & SCORING
# ═════════════════════════════════════════════════════════════════════════════

def parse_weights(raw: str = "") -> dict[str, float]:
    raw = raw.strip() or os.getenv("WEIGHTS", "").strip()
    if raw:
        try:
            vals = [float(x.strip()) for x in raw.split(",")]
            if len(vals) == len(DIMENSIONS):
                w = dict(zip(DIMENSIONS, vals))
                log.info(f"Custom weights: {w}")
                return w
        except (ValueError, TypeError):
            pass
    equal = 1.0 / len(DIMENSIONS)
    log.info("Equal weights across all 7 dimensions.")
    return {d: equal for d in DIMENSIONS}


def compute_score(scores: dict, weights: dict) -> float:
    total_w = sum(weights.values()) or 1
    weighted = sum(
        (scores.get(d, {}).get("score", 0)
         if isinstance(scores.get(d), dict) else 0) * weights[d]
        for d in DIMENSIONS
    )
    return round((weighted / (5 * total_w)) * 100, 1)


def priority_tier(score: float) -> str:
    if score >= 70: return "High"
    if score >= 45: return "Medium"
    return "Low"


def priority_display(score: float) -> str:
    return f"{priority_tier(score)} ({score})"


# ═════════════════════════════════════════════════════════════════════════════
# SITE DISCOVERY & CONTENT LOADING
# ═════════════════════════════════════════════════════════════════════════════

def find_site_folders(source_filter: str = "") -> list[Path]:
    if not SITES_DIR.exists():
        raise FileNotFoundError(
            f"sites/ not found at {SITES_DIR}. Run legacy_scraper.py first."
        )
    folders = []
    for p in sorted(SITES_DIR.iterdir()):
        if not p.is_dir():
            continue
        has_content = (
            (p / "pages_text.csv").exists() or
            (p / "posts_text.csv").exists()
        )
        if not has_content:
            continue
        if source_filter and source_filter.lower() not in p.name.lower():
            continue
        folders.append(p)
    if not folders:
        msg = "No site folders with content found"
        if source_filter:
            msg += f" matching '{source_filter}'"
        raise FileNotFoundError(msg + f" under {SITES_DIR}.")
    return folders


def source_name(folder: Path) -> str:
    raw = folder.name.split("_")[0]
    raw = re.sub(r"-(com|ai|world|io|net|org|co|gov|edu)$",
                 "", raw, flags=re.IGNORECASE)
    return raw.replace("-", " ").title()


def load_content(folder: Path, max_pages: int, ocr_min_len: int) -> str:
    chunks: list[str] = []
    for csv_name in ("pages_text.csv", "posts_text.csv"):
        csv_path = folder / csv_name
        if not csv_path.exists():
            continue
        try:
            df = pd.read_csv(csv_path, dtype=str).fillna("")
            # Sort by relevance_score descending if column exists
            if "relevance_score" in df.columns:
                df["relevance_score"] = pd.to_numeric(
                    df["relevance_score"], errors="coerce"
                ).fillna(0)
                df = df.sort_values("relevance_score", ascending=False)
            df = df.head(max_pages)
            for _, row in df.iterrows():
                url   = row.get("url", row.get("source_image", ""))
                title = row.get("page_title", row.get("source", ""))
                text  = row.get("clean_text", row.get("combined_text",
                        row.get("provided_text", "")))
                kw    = row.get("keyword_hits", "")
                score = row.get("relevance_score", "")
                if text.strip():
                    header = f"[PAGE: {title} | {url} | relevance={score} | keywords={kw}]"
                    chunks.append(f"{header}\n{text[:3000]}")
        except Exception as e:
            log.warning(f"Could not read {csv_name} in {folder.name}: {e}")

    ocr_path = folder / "ocr_output.csv"
    if ocr_path.exists():
        try:
            df = pd.read_csv(ocr_path, dtype=str).fillna("")
            for _, row in df.iterrows():
                ocr_text = row.get("extracted_text", "").strip()
                if len(ocr_text) >= ocr_min_len:
                    src = row.get("source_page_url", row.get("source_image", ""))
                    chunks.append(f"[IMAGE TEXT | {src}]\n{ocr_text}")
        except Exception as e:
            log.warning(f"Could not read ocr_output.csv in {folder.name}: {e}")

    return "\n\n---\n\n".join(chunks)


# ═════════════════════════════════════════════════════════════════════════════
# PROMPTS
# ═════════════════════════════════════════════════════════════════════════════

_EXTRACTION_SYSTEM = textwrap.dedent(f"""
    You are a competitive intelligence analyst for AmaliTech, an AI-first
    technology services company in Ghana and Rwanda serving European enterprise
    clients.

    Your focus: extract ALL services and products related to legacy system
    modernisation using AI. This specifically includes:

    1. COBOL / mainframe modernisation — migrating from COBOL, PL/I, JCL, IBM Z,
       AS/400 to modern languages (Java, Python, C#) and cloud-native architectures
    2. Java version migration — upgrading from Java 8, Java 11 to Java 17, Java 21
       or later LTS versions, including Spring Boot, Jakarta EE framework upgrades
    3. AI-assisted refactoring tools and platforms — watsonx Code Assistant,
       Amazon Q Code Transformation, GitHub Copilot App Modernization, OpenRewrite,
       Moderne, Swimm, Kodesage, Blu Age, Heirloom, etc.
    4. General legacy application modernisation services where AI is a key component
    5. Academic research, case studies, or proof points about effectiveness

    For each service or product found return a JSON object:
    {{
      "name": "short clear name",
      "source": "company or organisation name",
      "type": "service | product | tool | research",
      "category": "one of: {' | '.join(SERVICE_CATEGORIES)}",
      "customer_maturity": "one of: {' | '.join(MATURITY_LEVELS)}",
      "modernisation_focus": "one or more of: mainframe_cobol | java_migration | ai_refactoring | general_legacy",
      "description": "2-4 sentences — what it does, who it is for, how AI is used",
      "maturity_level": "experimental | emerging | established | proven",
      "evidence": "specific quote or fact from the content confirming this exists",
      "source_url": "URL where found or empty string",
      "academic_research": "any referenced studies, papers, or academic backing — or empty string"
    }}

    Return ONLY a valid JSON array. No markdown, no preamble.
    If nothing relevant found return: []
""").strip()

_EXTRACTION_USER = "Source: {name}\n\nContent:\n{content}"

_SCORING_SYSTEM = textwrap.dedent("""
    You are scoring legacy modernisation services against AmaliTech's priority
    matrix. AmaliTech is an AI-first tech services company delivering from
    Ghana + Rwanda to European enterprise clients.

    AmaliTech context for legacy modernisation specifically:
    - Strong Java, Python, cloud-native capability in Ghana/Rwanda
    - COBOL expertise is rare — but AI tools (watsonx, Amazon Q, OpenRewrite)
      reduce the need for deep COBOL knowledge, lowering the barrier
    - Key manufacturing accounts (Schaeffler, Knauf) likely have Java 8 estates
      that need upgrading; telco accounts (Telekom, 1&1) may have legacy systems
    - TISAX and ISO 27001 certified — trusted for sensitive enterprise codebases
    - AWS Advanced Partner — can leverage Amazon Q Code Transformation
    - European compliance positioning is a strong differentiator for data-sensitive
      mainframe modernisation work

    Score 1–5 per dimension. Ground every justification in AmaliTech's
    specific context. One sentence per justification.

    Dimensions:
    {dim_guide}

    Return ONLY valid JSON, no markdown:
    {{
      "market_impact":        {{"score": <1-5>, "justification": "<one sentence>"}},
      "effort":               {{"score": <1-5>, "justification": "<one sentence>"}},
      "scalability":          {{"score": <1-5>, "justification": "<one sentence>"}},
      "revenue_potential":    {{"score": <1-5>, "justification": "<one sentence>"}},
      "market_credibility":   {{"score": <1-5>, "justification": "<one sentence>"}},
      "talent_availability":  {{"score": <1-5>, "justification": "<one sentence>"}},
      "strategic_fit":        {{"score": <1-5>, "justification": "<one sentence>"}}
    }}
""").strip()

_SCORING_USER = (
    "Source: {source}\n"
    "Service/Product: {name}\n"
    "Type: {type}\n"
    "Category: {category}\n"
    "Modernisation Focus: {focus}\n"
    "Maturity Level: {maturity_level}\n"
    "Description: {description}\n\n"
    "Score against AmaliTech's priority matrix."
)

_BRIEF_SYSTEM = textwrap.dedent("""
    You are a research analyst writing a concise brief for AmaliTech's leadership
    on AI-assisted legacy system modernisation.

    AmaliTech context:
    - AI-first technology services company, Ghana + Rwanda delivery, European clients
    - Considering whether to offer legacy modernisation as a service
    - Key accounts: Schaeffler (manufacturing), Deutsche Telekom, Knauf, 1&1
    - Strong Java/Python/cloud capability; AWS Advanced Partner; TISAX certified

    You will be given scraped content from competitor websites, tool vendors,
    and research sources. Answer each question with:
    - A direct, evidence-based answer (3-6 sentences)
    - Specific named companies, tools, or studies where relevant
    - An explicit AmaliTech implication at the end of each answer

    Be concrete. Avoid vague statements. If evidence is thin, say so.
""").strip()

_BRIEF_USER = textwrap.dedent("""
    Based on the research content below, answer each of the four questions.

    Return a JSON object with this exact structure:
    {{
      "Q1": {{
        "question": "Who is doing AI-assisted legacy/mainframe/COBOL modernisation?",
        "answer": "<your answer>",
        "amalitech_implication": "<what this means for AmaliTech>"
      }},
      "Q2": {{
        "question": "Does it work — what is the maturity level?",
        "answer": "<your answer>",
        "amalitech_implication": "<what this means for AmaliTech>"
      }},
      "Q3": {{
        "question": "What state-of-the-art tools and approaches exist (including academic)?",
        "answer": "<your answer>",
        "amalitech_implication": "<what this means for AmaliTech>"
      }},
      "Q4": {{
        "question": "What exists for Java 8/11 → 17/21 migration using AI?",
        "answer": "<your answer>",
        "amalitech_implication": "<what this means for AmaliTech>"
      }}
    }}

    Return ONLY valid JSON. No markdown, no preamble.

    Research content:
    {content}
""").strip()


# ═════════════════════════════════════════════════════════════════════════════
# ANALYSIS FUNCTIONS
# ═════════════════════════════════════════════════════════════════════════════

def extract_services(name: str, content: str) -> list[dict]:
    log.info("  Extracting legacy modernisation services...")
    raw    = call_ai(
        _EXTRACTION_SYSTEM,
        _EXTRACTION_USER.format(name=name, content=content[:6_000]),
        max_tokens=4096,
    )
    result = _parse_json(raw, context=name)
    if not isinstance(result, list):
        log.warning(f"  Unexpected extraction response for {name}.")
        return []
    log.info(f"  {len(result)} service(s)/product(s) found.")
    return result


def score_service(source: str, service: dict) -> dict:
    raw = call_ai(
        _SCORING_SYSTEM.format(dim_guide=_DIM_GUIDE),
        _SCORING_USER.format(
            source=source,
            name=service.get("name", ""),
            type=service.get("type", ""),
            category=service.get("category", ""),
            focus=service.get("modernisation_focus", ""),
            maturity_level=service.get("maturity_level", ""),
            description=service.get("description", ""),
        ),
        max_tokens=1024,
    )
    result = _parse_json(raw, context=service.get("name", ""))
    return result if isinstance(result, dict) else {}


def generate_research_brief(all_content: str) -> dict:
    """
    Generate answers to the four pillar lead questions using all
    scraped content combined (up to 80k chars).
    """
    log.info("  Generating research brief (answering 4 pillar questions)...")
    raw = call_ai(
        _BRIEF_SYSTEM,
        _BRIEF_USER.format(content=all_content[:6_000]),
        max_tokens=4096,
    )
    result = _parse_json(raw, context="research_brief")
    return result if isinstance(result, dict) else {}


# ═════════════════════════════════════════════════════════════════════════════
# ROW BUILDER
# ═════════════════════════════════════════════════════════════════════════════

def build_rows(source: str, services: list[dict],
               all_scores: list[dict], weights: dict) -> list[dict]:
    rows = []
    for svc, scores in zip(services, all_scores):
        score = compute_score(scores, weights)
        row: dict = {
            "source":               source,
            "service_name":         svc.get("name", ""),
            "type":                 svc.get("type", ""),
            "category":             svc.get("category", ""),
            "customer_maturity":    svc.get("customer_maturity", ""),
            "modernisation_focus":  svc.get("modernisation_focus", ""),
            "maturity_level":       svc.get("maturity_level", ""),
            "ai_classification":    "core_ai",
            "description":          svc.get("description", ""),
            "evidence":             svc.get("evidence", ""),
            "academic_research":    svc.get("academic_research", ""),
            "source_url":           svc.get("source_url", ""),
            "priority_score":       score,
            "priority_tier":        priority_tier(score),
            "priority_display":     priority_display(score),
        }
        for dim in DIMENSIONS:
            d = scores.get(dim, {})
            row[f"{dim}_score"]         = d.get("score", "") if isinstance(d, dict) else ""
            row[f"{dim}_justification"] = d.get("justification", "") if isinstance(d, dict) else ""
        rows.append(row)

    rows.sort(key=lambda r: r["priority_score"], reverse=True)
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═════════════════════════════════════════════════════════════════════════════

def _write_long_list_sheet(ws, rows: list[dict], title: str) -> None:
    """Write Initiative Long List sheet — competitor + service columns."""
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = title
    c.font      = _font(bold=True, color=WHITE, size=13)
    c.fill      = _fill(NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers    = ["#", "Source / Competitor", "AI Service Category",
                  "Customer Maturity", "Service / Product",
                  "Modernisation Focus", "Priority"]
    col_widths = [5, 22, 28, 24, 32, 24, 18]

    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=2, column=col_i, value=h)
        cell.font      = _font(bold=True, color=WHITE, size=10)
        cell.fill      = _fill(MID_BLUE)
        cell.alignment = _align("center")
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for row_num, row in enumerate(rows, start=1):
        excel_row = row_num + 2
        tier = row.get("priority_tier", "")
        alt  = row_num % 2 == 0
        bg   = _fill(LIGHTER if alt else WHITE)

        values = [
            row_num,
            row.get("source", ""),
            row.get("category", ""),
            row.get("customer_maturity", ""),
            row.get("service_name", ""),
            row.get("modernisation_focus", ""),
            row.get("priority_display", ""),
        ]

        for col_i, value in enumerate(values, start=1):
            cell           = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border    = BORDER
            cell.font      = _font(size=9)
            cell.alignment = _align(
                "center" if col_i in (1, 4, 7) else "left"
            )
            cell.fill = bg

        p_cell = ws.cell(row=excel_row, column=7)
        if tier == "High":
            p_cell.fill = _fill(HIGH_BG)
            p_cell.font = _font(bold=True, color=HIGH_FG, size=9)
        elif tier == "Medium":
            p_cell.fill = _fill(MED_BG)
            p_cell.font = _font(bold=True, color=MED_FG, size=9)
        else:
            p_cell.fill = _fill(LOW_BG)
            p_cell.font = _font(bold=True, color=LOW_FG, size=9)

        ws.row_dimensions[excel_row].height = 28


def write_long_list_xlsx(all_rows: list[dict], path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    by_source: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        by_source[row["source"]].append(row)

    # Combined sheet first
    ws_all = wb.create_sheet(title="All Sources")
    combined = sorted(
        all_rows,
        key=lambda r: (r.get("source", ""), -float(r.get("priority_score", 0))),
    )
    _write_long_list_sheet(
        ws_all, combined,
        "Legacy Modernisation — Initiative Long List (All Sources)"
    )

    # Per-source sheets
    for src in sorted(by_source.keys()):
        rows = sorted(
            by_source[src],
            key=lambda r: -float(r.get("priority_score", 0)),
        )
        ws = wb.create_sheet(title=src[:28])
        _write_long_list_sheet(ws, rows, f"Legacy Modernisation — {src}")

    wb.save(path)
    log.info(f"  Long list saved: {path.name}  "
             f"({len(all_rows)} items, {len(by_source)} sources)")


def write_detailed_workbook(source: str, rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scored Services"

    total_cols = 11 + len(DIMENSIONS) * 2
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    t = ws["A1"]
    t.value     = f"Legacy Modernisation Analysis — {source}"
    t.font      = _font(bold=True, color=WHITE, size=13)
    t.fill      = _fill(NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    base_headers = [
        "Service / Product", "Source", "Type", "Category",
        "Customer Maturity", "Modernisation Focus", "Maturity Level",
        "Description", "Evidence", "Academic Research",
        "Priority Score", "Priority Tier",
    ]
    # adjust col_widths to match base_headers count (12) + dims
    base_widths = [30, 20, 12, 26, 20, 22, 14, 45, 35, 30, 14, 12]
    dim_headers = []
    dim_widths  = []
    for d in DIMENSIONS:
        dim_headers += [f"{DIM_LABELS[d]} Score", f"{DIM_LABELS[d]} Justification"]
        dim_widths  += [10, 36]

    all_headers = base_headers + dim_headers
    all_widths  = base_widths  + dim_widths

    for col_i, (h, w) in enumerate(zip(all_headers, all_widths), start=1):
        cell           = ws.cell(row=2, column=col_i, value=h)
        cell.font      = _font(bold=True, color=WHITE, size=9)
        cell.fill      = _fill(MID_BLUE)
        cell.alignment = _align("center", wrap=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    for row_num, row in enumerate(rows, start=1):
        excel_row = row_num + 2
        tier = row.get("priority_tier", "")
        alt  = row_num % 2 == 0

        base_values = [
            row.get("service_name", ""),
            row.get("source", ""),
            row.get("type", ""),
            row.get("category", ""),
            row.get("customer_maturity", ""),
            row.get("modernisation_focus", ""),
            row.get("maturity_level", ""),
            row.get("description", ""),
            row.get("evidence", ""),
            row.get("academic_research", ""),
            row.get("priority_score", ""),
            row.get("priority_tier", ""),
        ]
        dim_values = []
        for d in DIMENSIONS:
            dim_values += [
                row.get(f"{d}_score", ""),
                row.get(f"{d}_justification", ""),
            ]

        for col_i, value in enumerate(base_values + dim_values, start=1):
            if isinstance(value, (list, dict)):
                value = json.dumps(value)
            cell           = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border    = BORDER
            cell.alignment = _align("left", wrap=True)
            cell.font      = _font(size=9)
            cell.fill      = _fill(LIGHTER if alt else WHITE)

        for col_i in (11, 12):
            cell = ws.cell(row=excel_row, column=col_i)
            if tier == "High":
                cell.fill = _fill(HIGH_BG)
                cell.font = _font(bold=True, color=HIGH_FG, size=9)
            elif tier == "Medium":
                cell.fill = _fill(MED_BG)
                cell.font = _font(bold=True, color=MED_FG, size=9)
            else:
                cell.fill = _fill(LOW_BG)
                cell.font = _font(bold=True, color=LOW_FG, size=9)

        ws.row_dimensions[excel_row].height = 45

    wb.save(path)
    log.info(f"  Detailed workbook: {path.name}")


# ═════════════════════════════════════════════════════════════════════════════
# MARKDOWN RESEARCH BRIEF
# ═════════════════════════════════════════════════════════════════════════════

def write_research_brief_md(
    brief: dict,
    all_rows: list[dict],
    path: Path,
    backend: str,
) -> None:
    """
    Write a Markdown research brief that answers the 4 pillar lead questions
    followed by the full scored service list as a reference table.
    """
    ts = datetime.now().strftime("%B %Y")
    lines = [
        "# Legacy Modernisation Research Brief",
        "",
        f"**Prepared for:** AI Engineering & Automation Pillar Lead  ",
        f"**Prepared by:** Benchmarking Team — AmaliTech  ",
        f"**Date:** {ts}  ",
        f"**AI backend used:** {backend.upper()}  ",
        "",
        "---",
        "",
        "## Executive Summary",
        "",
        "This brief answers four strategic questions about AI-assisted legacy "
        "system modernisation — covering COBOL/mainframe migration and Java "
        "version upgrades — to help AmaliTech assess whether to offer this as "
        "a service.",
        "",
        "---",
        "",
    ]

    q_labels = {
        "Q1": "Who is doing this?",
        "Q2": "Does it work — what is the maturity?",
        "Q3": "State-of-the-art tools and approaches",
        "Q4": "Java 8/11 → 17/21 migration using AI",
    }

    for q_id in ["Q1", "Q2", "Q3", "Q4"]:
        q_data = brief.get(q_id, {})
        label  = q_labels.get(q_id, q_id)
        lines += [
            f"## {q_id}: {label}",
            "",
            f"**Question:** {q_data.get('question', '')}",
            "",
            "**Answer:**",
            "",
            q_data.get("answer", "_No answer generated — check API response._"),
            "",
            "**AmaliTech implication:**",
            "",
            f"> {q_data.get('amalitech_implication', '_Not available._')}",
            "",
            "---",
            "",
        ]

    # ── Scored service list ──
    lines += [
        "## Full Scored Service & Product List",
        "",
        "Sorted by priority score (High → Medium → Low). "
        "Source column shows which company or platform the service/product was found on.",
        "",
    ]

    # Group by source
    by_source: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        by_source[row["source"]].append(row)

    global_num = 1
    for src in sorted(by_source.keys()):
        rows = sorted(
            by_source[src],
            key=lambda r: -float(r.get("priority_score", 0)),
        )
        lines.append(f"### {src}")
        lines.append("")
        lines.append(
            "| # | Service / Product | Focus | Maturity | Priority |"
        )
        lines.append("|---|---|---|---|---|")
        for row in rows:
            lines.append(
                f"| {global_num} "
                f"| {row.get('service_name', '')} "
                f"| {row.get('modernisation_focus', '')} "
                f"| {row.get('maturity_level', '')} "
                f"| {row.get('priority_display', '')} |"
            )
            global_num += 1
        lines.append("")

    lines += [
        "---",
        "",
        f"*Research based on scraped website data — {ts}. "
        f"Analysis generated using {backend.upper()}.*",
        "",
    ]

    path.write_text("\n".join(lines), encoding="utf-8")
    log.info(f"  Research brief: {path.name}")


# ═════════════════════════════════════════════════════════════════════════════
# MERGE EXISTING
# ═════════════════════════════════════════════════════════════════════════════

def load_existing() -> pd.DataFrame:
    if MASTER_CSV.exists():
        try:
            df = pd.read_csv(MASTER_CSV, dtype=str).fillna("")
            log.info(f"Loaded {len(df)} existing rows from master CSV.")
            return df
        except Exception as e:
            log.warning(f"Could not read master CSV: {e}")
    return pd.DataFrame()


def merge_results(existing: pd.DataFrame, new_rows: list[dict]) -> pd.DataFrame:
    new_df = pd.DataFrame(new_rows) if new_rows else pd.DataFrame()
    if existing.empty:
        return new_df
    if new_df.empty:
        return existing
    new_sources = new_df["source"].unique().tolist()
    existing_filtered = existing[~existing["source"].isin(new_sources)]
    merged = pd.concat([existing_filtered, new_df], ignore_index=True)
    return merged.sort_values(
        ["source", "priority_score"], ascending=[True, False]
    ).reset_index(drop=True)


# ═════════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═════════════════════════════════════════════════════════════════════════════

def run(
    source_filter: str = "",
    weights_str: str = "",
    max_pages: int = 0,
    ocr_min_len: int = 0,
    dry_run: bool = False,
    rerun_all: bool = False,
) -> pd.DataFrame:

    rerun_all   = rerun_all or os.getenv("RERUN_ALL", "").strip() in ("1", "true", "yes")
    max_pages   = max_pages   or int(os.getenv("MAX_PAGES_PER_SITE", 40))
    ocr_min_len = ocr_min_len or int(os.getenv("OCR_MIN_LENGTH", 30))
    weights     = parse_weights(weights_str)

    OUTPUT_DIR.mkdir(exist_ok=True)

    if rerun_all:
        clear_state()
    state = load_state()

    all_folders = find_site_folders(source_filter)
    pending     = [f for f in all_folders if f.name not in state]
    completed   = [f for f in all_folders if f.name in state]

    log.info(f"\nAI backend       : {AI_BACKEND.upper()}"
             + (f" ({GROQ_MODEL})" if AI_BACKEND == "groq" else f" ({CLAUDE_MODEL})"))
    log.info(f"Site folders     : {len(all_folders)}")
    log.info(f"  Already done   : {len(completed)}")
    log.info(f"  Pending        : {len(pending)}")

    if completed:
        log.info("  Completed:")
        for f in completed:
            info = state[f.name]
            log.info(f"    ✓ {source_name(f):<28} "
                     f"({info.get('services', '?')} services)")
    if pending:
        log.info("  Pending:")
        for f in pending:
            log.info(f"    … {source_name(f)}")

    if dry_run:
        log.info("\nDRY RUN — no API calls made.")
        return pd.DataFrame()

    if not pending:
        log.info("\nNo new folders to process. Rebuilding outputs...")
        existing = load_existing()
        if not existing.empty:
            _write_outputs(existing, state, weights)
        return existing

    # ── Collect all content for the research brief (across all sources) ──
    # We accumulate from both existing and new sources
    all_content_chunks: list[str] = []

    # Load content from already-completed folders for the brief
    for folder in completed:
        chunk = load_content(folder, max_pages, ocr_min_len)
        if chunk.strip():
            all_content_chunks.append(
                f"=== {source_name(folder)} ===\n{chunk[:8000]}"
            )

    new_rows: list[dict] = []

    for folder in pending:
        name = source_name(folder)
        log.info(f"\n{'─'*60}")
        log.info(f"  {name}  ({folder.name})")
        log.info(f"{'─'*60}")

        content = load_content(folder, max_pages, ocr_min_len)
        if not content.strip():
            log.warning("  No content — skipping.")
            state[folder.name] = {
                "processed_at": datetime.now().isoformat(timespec="seconds"),
                "services": 0, "source": name, "skipped": True,
            }
            save_state(state)
            continue

        log.info(f"  Content: ~{len(content):,} chars")
        all_content_chunks.append(f"=== {name} ===\n{content[:8000]}")

        # ── Extract ──
        services = extract_services(name, content)
        if not services:
            log.warning("  No services extracted — skipping.")
            state[folder.name] = {
                "processed_at": datetime.now().isoformat(timespec="seconds"),
                "services": 0, "source": name, "skipped": True,
            }
            save_state(state)
            continue

        # ── Score ──
        scored: list[dict] = []
        for i, svc in enumerate(services, 1):
            log.info(f"  Scoring [{i}/{len(services)}]: {svc.get('name', '?')}")
            scores = score_service(name, svc)
            scored.append(scores)
            time.sleep(0.3)

        rows = build_rows(name, services, scored, weights)
        new_rows.extend(rows)

        # Per-source detailed xlsx
        safe        = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
        detail_path = OUTPUT_DIR / f"{safe}_legacy_scored.xlsx"
        write_detailed_workbook(name, rows, detail_path)

        state[folder.name] = {
            "processed_at": datetime.now().isoformat(timespec="seconds"),
            "services": len(rows), "source": name, "skipped": False,
        }
        save_state(state)

        # ── Flush rows to master CSV immediately (crash-safe) ──
        existing_now = load_existing()
        partial      = merge_results(existing_now, rows)
        OUTPUT_DIR.mkdir(exist_ok=True)
        partial.to_csv(MASTER_CSV, index=False, encoding="utf-8")

        log.info(f"  ✓ {name} — {len(rows)} items scored.")

    # ── Merge + write outputs ──
    existing = load_existing()
    master   = merge_results(existing, new_rows)

    if master.empty:
        log.warning("No results to write.")
        return master

    # ── Generate research brief from all available content ──
    combined_content = "\n\n".join(all_content_chunks)
    brief = generate_research_brief(combined_content)

    _write_outputs(master, state, weights, brief=brief)
    return master


def _write_outputs(
    master: pd.DataFrame,
    state: dict,
    weights: dict = None,
    brief: dict = None,
) -> None:
    master.to_csv(MASTER_CSV, index=False, encoding="utf-8")
    log.info(f"  Master CSV: {MASTER_CSV.name}  ({len(master)} rows)")

    ts       = datetime.now().strftime("%Y%m%d")
    all_rows = master.to_dict("records")

    # Long list xlsx
    ll_path = OUTPUT_DIR / f"{ts}_legacy_long_list.xlsx"
    write_long_list_xlsx(all_rows, ll_path)

    # Research brief markdown
    if brief:
        brief_path = OUTPUT_DIR / f"{ts}_legacy_research_brief.md"
        write_research_brief_md(brief, all_rows, brief_path, AI_BACKEND)

    done = sum(1 for v in state.values() if not v.get("skipped", False))
    log.info(f"\n{'='*60}")
    log.info(f"  Sources processed : {done} / {len(state)} tracked")
    log.info(f"  Total items       : {len(master)}")
    for tier in ["High", "Medium", "Low"]:
        n = (master["priority_tier"] == tier).sum()
        log.info(f"  {tier:<8} priority  : {n}")
    log.info(f"  Outputs in        : {OUTPUT_DIR}/")
    log.info(f"    {ts}_legacy_long_list.xlsx")
    if brief:
        log.info(f"    {ts}_legacy_research_brief.md")
    log.info(f"    {MASTER_CSV.name}")
    log.info(f"{'='*60}\n")


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="AmaliTech Legacy Modernisation Analysis"
    )
    parser.add_argument("--source", "-s", default="",
                        help="Filter to a single source (e.g. 'capgemini')")
    parser.add_argument("--weights", "-w", default="",
                        help="Comma-separated weights for 7 dimensions")
    parser.add_argument("--max-pages", "-p", type=int, default=0,
                        help="Max pages per source (default: 40)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show pending sources, no API calls")
    parser.add_argument("--rerun-all", action="store_true",
                        help="Ignore state, reprocess everything "
                             "(also via RERUN_ALL=1)")
    args = parser.parse_args()

    run(
        source_filter=args.source,
        weights_str=args.weights,
        max_pages=args.max_pages,
        dry_run=args.dry_run,
        rerun_all=args.rerun_all,
    )


if __name__ == "__main__":
    main()
