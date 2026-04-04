"""lib/pipelines/news_monitoring.py — extraction, scoring, alerting, and Excel output."""
from __future__ import annotations

import json
import os
import re
import textwrap
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

from lib.core import call_ai, parse_json, TIER_CONTENT, AI_BACKEND
from lib.keywords.news_monitoring import ALERT_KEYWORDS
import logging
log = logging.getLogger("analyse")

DIMENSIONS: list = []  # no scoring matrix — priority is rule+AI hybrid

# ── ALERT TRIGGERS ────────────────────────────────────────────────────────────
ALERT_CATEGORIES  = {"security", "critical update", "vulnerability", "zero day"}
ALERT_INNOVATION  = True   # innovation_signal = yes always triggers alert

# ── EXTRACTION PROMPT ─────────────────────────────────────────────────────────
_EXTRACTION_SYSTEM = textwrap.dedent("""
    You are a technical intelligence analyst monitoring news and tech blogs for
    AmaliTech — an AI-first technology services company serving European enterprise
    clients in manufacturing, telecoms, and financial services.

    From the content provided, extract each distinct article or news item.

    For each item return a JSON object with EXACTLY these keys:
    {
      "title": "article title",
      "source": "publication or blog name",
      "url": "direct article URL if found, else the page URL",
      "category": "one of: AI/ML | DevOps | Platform Engineering | Security | Cloud | Data Engineering | MLOps | Tools | Concepts | Other",
      "tools": "comma-separated tools/technologies mentioned (e.g. Kubernetes, ArgoCD, Terraform)",
      "cloud": "comma-separated cloud platforms mentioned (e.g. AWS, Azure, GCP)",
      "concepts": "comma-separated concepts mentioned (e.g. GitOps, IaC, Zero Trust)",
      "summary": "2-3 sentences — what this article is about and why it matters",
      "impact": "1 sentence — practical implication for a tech services team",
      "innovation_signal": "yes or no — does this describe a new approach, pattern, or tool that changes how things are done?",
      "innovation_detail": "if innovation_signal is yes: brief explanation of what is new, else empty string"
    }

    Return ONLY a valid JSON array. No markdown, no preamble.
    Extract every distinct article you can find. If no articles found return: []
""").strip()


def extract_services(name: str, content: str) -> list[dict]:
    log.info("  Extracting news articles...")
    raw    = call_ai(_EXTRACTION_SYSTEM, f"Source: {name}\n\nContent:\n{content[:TIER_CONTENT]}")
    result = parse_json(raw, context=name)
    if not isinstance(result, list):
        log.warning(f"  Unexpected extraction response for {name}.")
        return []
    log.info(f"  {len(result)} article(s) found.")
    return result


# ── PRIORITY SCORING ──────────────────────────────────────────────────────────
def _score_priority(article: dict) -> tuple[int, str]:
    score = 0
    cat   = article.get("category", "").lower()
    tools = article.get("tools", "").lower()
    cloud = article.get("cloud", "").lower()
    summ  = (article.get("summary", "") + " " + article.get("title", "")).lower()

    if "security" in cat or "vulnerability" in cat:  score += 3
    if "ai/ml" in cat or "mlops" in cat:             score += 2
    if "kubernetes" in tools or "k8s" in tools:      score += 2
    if "aws" in cloud or "azure" in cloud:           score += 1
    if article.get("innovation_signal", "").lower() == "yes": score += 2
    if any(kw in summ for kw in ALERT_KEYWORDS):     score += 3

    tier = "high" if score >= 5 else ("medium" if score >= 3 else "low")
    return score, tier


def _should_alert(article: dict) -> bool:
    cat  = article.get("category", "").lower()
    summ = (article.get("summary","") + " " + article.get("title","")).lower()
    if article.get("priority") == "high":
        return True
    if any(ac in cat for ac in ALERT_CATEGORIES):
        return True
    if ALERT_INNOVATION and article.get("innovation_signal","").lower() == "yes":
        return True
    if any(kw in summ for kw in ALERT_KEYWORDS):
        return True
    return False


# ── ROW BUILDER ───────────────────────────────────────────────────────────────
def build_rows(source: str, services: list[dict]) -> list[dict]:
    rows = []
    ts   = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    run  = datetime.utcnow().strftime("%H:%M")
    for i, art in enumerate(services, start=1):
        score, tier = _score_priority(art)
        alert       = _should_alert({**art, "priority": tier})
        rows.append({
            "source":           source,
            "timestamp":        ts,
            "run":              run,
            "title":            art.get("title", ""),
            "article_source":   art.get("source", source),
            "url":              art.get("url", ""),
            "category":         art.get("category", ""),
            "tools":            art.get("tools", ""),
            "cloud":            art.get("cloud", ""),
            "concepts":         art.get("concepts", ""),
            "summary":          art.get("summary", ""),
            "impact":           art.get("impact", ""),
            "innovation_signal":art.get("innovation_signal", "no"),
            "innovation_detail":art.get("innovation_detail", ""),
            "priority_score":   score,
            "priority_tier":    tier,
            "alert_triggered":  "yes" if alert else "no",
        })
    rows.sort(key=lambda r: -r["priority_score"])
    return rows


# ── DEDUPLICATION ─────────────────────────────────────────────────────────────
def load_seen_urls(output_dir: Path) -> set[str]:
    f = output_dir / "news_seen_urls.json"
    if f.exists():
        try:
            return set(json.loads(f.read_text(encoding="utf-8")))
        except Exception:
            pass
    return set()


def save_seen_urls(output_dir: Path, seen: set[str]) -> None:
    f = output_dir / "news_seen_urls.json"
    f.write_text(json.dumps(sorted(seen), indent=2, ensure_ascii=False), encoding="utf-8")


def deduplicate(rows: list[dict], seen: set[str]) -> tuple[list[dict], set[str]]:
    new_rows, new_seen = [], set()
    for row in rows:
        url = row.get("url", "").strip()
        if url and url in seen:
            continue
        new_rows.append(row)
        if url:
            new_seen.add(url)
    return new_rows, new_seen


# ── SLACK ALERTING ────────────────────────────────────────────────────────────
def _post_slack(message: str) -> None:
    webhook = os.getenv("SLACK_WEBHOOK_URL", "").strip()
    if not webhook:
        log.info(f"[ALERT] {message}")
        return
    try:
        import requests as req
        resp = req.post(webhook, json={"text": message}, timeout=10)
        if resp.status_code != 200:
            log.warning(f"Slack webhook returned {resp.status_code}: {resp.text[:100]}")
    except Exception as e:
        log.warning(f"Slack alert failed: {e}")


def send_run_start_alert(run_time: str) -> None:
    _post_slack(f":rocket: *News monitoring run started* at {run_time} UTC")


def send_run_end_alert(scanned: int, relevant: int, high: int, alerts: int) -> None:
    _post_slack(
        f":white_check_mark: *News monitoring run completed*\n"
        f"• Articles scanned: {scanned}\n"
        f"• Relevant: {relevant}\n"
        f"• High priority: {high}\n"
        f"• Alerts fired: {alerts}"
    )


def send_article_alert(row: dict) -> None:
    icon = ":fire:" if row["priority_tier"] == "high" else ":bell:"
    innov = f"\n:bulb: *Innovation:* {row['innovation_detail']}" if row.get("innovation_detail") else ""
    msg = (
        f"{icon} *{row['priority_tier'].upper()} PRIORITY — {row['category']}*\n"
        f"*{row['title']}*\n"
        f"{row['summary']}\n"
        f":zap: *Impact:* {row['impact']}"
        f"{innov}\n"
        f":link: {row['url']}"
    )
    _post_slack(msg)


# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────
_COLUMNS = [
    "timestamp", "run", "title", "article_source", "url", "category",
    "tools", "cloud", "concepts", "summary", "impact",
    "innovation_signal", "innovation_detail",
    "priority_score", "priority_tier", "alert_triggered",
]
_WIDTHS = [18, 8, 40, 20, 50, 18, 30, 20, 30, 60, 40, 12, 40, 12, 10, 12]

_TIER_COLOURS = {
    "high":   ("C6EFCE", "276221"),
    "medium": ("FFEB9C", "9C6500"),
    "low":    ("FFFFFF", "000000"),
}


def _daily_xlsx_path(output_dir: Path) -> Path:
    return output_dir / f"news_{date.today().strftime('%Y%m%d')}.xlsx"


def append_to_daily_xlsx(rows: list[dict], output_dir: Path) -> Path:
    """Append rows to today's Excel file, creating it with headers if new."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY    = "1F3864"; MID_BLUE = "2E5DA8"; LIGHTER = "EBF3FA"
    _thin   = Side(style="thin", color="BFBFBF")
    BORDER  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=9):
        return Font(bold=bold, color=color, size=size, name="Calibri")
    def _align(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    path = _daily_xlsx_path(output_dir)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook(); ws = wb.active; ws.title = "News"
        # Title row
        ws.merge_cells(f"A1:{get_column_letter(len(_COLUMNS))}1")
        c = ws["A1"]
        c.value = f"AmaliTech News Intelligence — {date.today().strftime('%B %d, %Y')}"
        c.font  = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
        c.fill  = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        # Header row
        for col_i, (h, w) in enumerate(zip(_COLUMNS, _WIDTHS), start=1):
            cell = ws.cell(row=2, column=col_i, value=h.replace("_", " ").title())
            cell.font      = _font(bold=True, color="FFFFFF")
            cell.fill      = _fill(MID_BLUE)
            cell.alignment = _align("center")
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_i)].width = w
        ws.row_dimensions[2].height = 20
        ws.freeze_panes = "A3"
        start_row = 3

    for row_num, row in enumerate(rows, start=start_row):
        tier = row.get("priority_tier", "low")
        bg, fg = _TIER_COLOURS.get(tier, ("FFFFFF", "000000"))
        alt_bg = "EBF3FA" if row_num % 2 == 0 else "FFFFFF"
        for col_i, col in enumerate(_COLUMNS, start=1):
            val  = row.get(col, "")
            cell = ws.cell(row=row_num, column=col_i, value=val)
            cell.border    = BORDER
            cell.font      = _font(size=9)
            cell.alignment = _align(wrap=col in ("summary","impact","innovation_detail","title","url"))
            cell.fill      = _fill(alt_bg)
        # Colour priority and alert columns
        for col_name, col_i in [("priority_tier", _COLUMNS.index("priority_tier")+1),
                                  ("alert_triggered", _COLUMNS.index("alert_triggered")+1)]:
            cell = ws.cell(row=row_num, column=col_i)
            if col_name == "priority_tier":
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=fg, size=9)
            elif row.get("alert_triggered") == "yes":
                cell.fill = _fill("FFC7CE")
                cell.font = _font(bold=True, color="9C0006", size=9)
        ws.row_dimensions[row_num].height = 45

    wb.save(path)
    log.info(f"  Excel updated: {path.name}  (+{len(rows)} rows, total={ws.max_row - 2})")
    return path


# ── BRIEF / SUMMARY (required by analyse_new.py interface) ───────────────────
def generate_brief(all_content: str) -> dict:
    return {}  # news pipeline doesn't generate a cross-source brief


def write_brief_md(brief: dict, all_rows: list[dict], path: Path) -> None:
    """Write a run summary markdown — numbered list with priority, keywords, summary, link."""
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    lines = [
        f"# News Intelligence Run — {ts}",
        f"**AI backend:** {AI_BACKEND.upper()}  ",
        f"**Articles:** {len(all_rows)}  ",
        f"**High priority:** {sum(1 for r in all_rows if r['priority_tier']=='high')}  ",
        f"**Alerts fired:** {sum(1 for r in all_rows if r['alert_triggered']=='yes')}",
        "",
        "---",
        "",
    ]
    by_tier = {"high": [], "medium": [], "low": []}
    for row in all_rows:
        by_tier.get(row["priority_tier"], by_tier["low"]).append(row)

    n = 1
    for tier, label, icon in [("high","HIGH PRIORITY",":fire:"), ("medium","MEDIUM",""), ("low","LOW","")]:
        items = by_tier[tier]
        if not items:
            continue
        lines += [f"## {icon} {label} ({len(items)})", ""]
        for row in items:
            kws = ", ".join(filter(None, [row.get("tools",""), row.get("cloud",""), row.get("concepts","")]))
            innov = f"\n   > :bulb: **Innovation:** {row['innovation_detail']}" if row.get("innovation_detail") else ""
            alert = " 🔔" if row["alert_triggered"] == "yes" else ""
            lines += [
                f"**{n}.{alert} {row['title']}**  ",
                f"*{row['category']} | {row['article_source']}*  ",
                f"**Keywords:** {kws or '—'}  ",
                f"**Summary:** {row['summary']}  ",
                f"**Impact:** {row['impact']}{innov}  ",
                f"**Link:** {row['url']}",
                "",
            ]
            n += 1
    lines += ["---", f"*Generated {ts}*", ""]
    path.write_text("\n".join(lines), encoding="utf-8")
    log.info(f"  Run summary: {path.name}")
