"""lib/excel.py — shared openpyxl styling helpers used by all three pipelines."""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY    = "1F3864"; MID_BLUE = "2E5DA8"; LIGHTER = "EBF3FA"; WHITE = "FFFFFF"
HIGH_BG = "C6EFCE"; HIGH_FG = "276221"
MED_BG  = "FFEB9C"; MED_FG  = "9C6500"
LOW_BG  = "FFC7CE"; LOW_FG  = "9C0006"
GRAY    = "BFBFBF"

_thin  = Side(style="thin", color=GRAY)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def fill(h):  return PatternFill("solid", fgColor=h)
def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def title_row(ws, row_num: int, title: str, ncols: int, height: int = 28) -> None:
    ws.merge_cells(f"A{row_num}:{get_column_letter(ncols)}{row_num}")
    c = ws.cell(row=row_num, column=1, value=title)
    c.font      = font(bold=True, color=WHITE, size=13)
    c.fill      = fill(NAVY)
    c.alignment = align("center")
    ws.row_dimensions[row_num].height = height


def header_row(ws, row_num: int, headers: list[str], widths: list[int]) -> None:
    for col_i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row_num, column=col_i, value=h)
        cell.font      = font(bold=True, color=WHITE, size=9)
        cell.fill      = fill(MID_BLUE)
        cell.alignment = align("center", wrap=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[row_num].height = 30


def priority_cell(cell, tier: str) -> None:
    if tier == "High":
        cell.fill = fill(HIGH_BG); cell.font = font(bold=True, color=HIGH_FG, size=9)
    elif tier == "Medium":
        cell.fill = fill(MED_BG);  cell.font = font(bold=True, color=MED_FG,  size=9)
    else:
        cell.fill = fill(LOW_BG);  cell.font = font(bold=True, color=LOW_FG,  size=9)


def write_row(ws, excel_row: int, values: list, row_num: int,
              center_cols: set = None, priority_col: int = None,
              priority_tier: str = "", row_height: int = 40) -> None:
    alt = row_num % 2 == 0
    bg  = fill(LIGHTER if alt else WHITE)
    center_cols = center_cols or set()
    for col_i, val in enumerate(values, start=1):
        if isinstance(val, (list, dict)):
            import json
            val = json.dumps(val)
        cell = ws.cell(row=excel_row, column=col_i, value=val)
        cell.border    = BORDER
        cell.font      = font(size=9)
        cell.alignment = align("center" if col_i in center_cols else "left", wrap=True)
        cell.fill      = bg
    if priority_col:
        priority_cell(ws.cell(row=excel_row, column=priority_col), priority_tier)
    ws.row_dimensions[excel_row].height = row_height
