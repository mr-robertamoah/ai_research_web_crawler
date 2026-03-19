"""
AmaliTech Competitor AI Services — Analysis Script
====================================================
Incremental by default: tracks which site folders have already been analysed
in output/processed_folders.json and only processes new ones each run.
Safe to run repeatedly while scraping is still in progress.

After each run, merges ALL results (existing + new) into one updated
Initiative Long List xlsx ready for the SharePoint Excel sheet.

HOW TO RUN:
  python analyse.py                          # process only NEW site folders (default)
  python analyse.py --competitor andela      # process only andela if not yet done
  python analyse.py --max-pages 10           # faster / lower API cost
  python analyse.py --weights 2,1,1,2,1,1,1.5
  python analyse.py --dry-run                # show pending folders, no API calls

ENVIRONMENT VARIABLES:
  ANTHROPIC_API_KEY     required
  RERUN_ALL             set to 1 to ignore state and reprocess everything from scratch
  WEIGHTS               comma-separated weights for 7 dimensions (default: equal)
  MAX_PAGES_PER_SITE    max pages to feed per competitor (default: 40)
  OCR_MIN_LENGTH        min chars for OCR row to be included (default: 30)

WEIGHT ORDER:
  market_impact, effort, scalability, revenue_potential,
  market_credibility, talent_availability, strategic_fit

EFFORT IS INVERSE: score 5 = low effort (fast to launch), score 1 = very high effort.

CUSTOMER MATURITY LEVELS (AmaliTech-defined):
  AI Explorer               — client just starting, no AI in place yet
  AI Practitioner           — client has some AI deployed, actively scaling it
  AI Champion               — client is AI-mature, needs advanced services
  AI Explorer → AI Practitioner
  AI Practitioner → AI Champion
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import time
import textwrap
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import anthropic
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("analyse")

# ── PATHS ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR      = Path(__file__).parent.resolve()
SITES_DIR       = SCRIPT_DIR / "sites"
OUTPUT_DIR      = SCRIPT_DIR / "output"
STATE_FILE      = OUTPUT_DIR / "processed_folders.json"   # tracks what's done
MASTER_CSV      = OUTPUT_DIR / "all_competitors_priority.csv"

# ── MODEL ─────────────────────────────────────────────────────────────────────
MODEL = "claude-sonnet-4-20250514"

# ── AMALITECH SERVICE CATEGORIES ─────────────────────────────────────────────
SERVICE_CATEGORIES = [
    "AI Advisory & Readiness",
    "AI Engineering & Automation",
    "AI Platforms & Agents",
    "AI-powered Solutions & New Revenue Models",
    "Talent & Staffing",
    "Other",
]

# ── CUSTOMER MATURITY LEVELS ─────────────────────────────────────────────────
MATURITY_LEVELS = [
    "AI Explorer",
    "AI Practitioner",
    "AI Champion",
    "AI Explorer → AI Practitioner",
    "AI Practitioner → AI Champion",
]

# ── PRIORITY MATRIX DIMENSIONS ────────────────────────────────────────────────
DIMENSIONS = [
    "market_impact",
    "effort",               # INVERSE — score 5 = low effort
    "scalability",
    "revenue_potential",
    "market_credibility",
    "talent_availability",
    "strategic_fit",
]

DIM_LABELS = {
    "market_impact":       "Market Impact",
    "effort":              "Effort (inverse: 5=low effort)",
    "scalability":         "Scalability",
    "revenue_potential":   "Revenue Potential",
    "market_credibility":  "Market Credibility",
    "talent_availability": "Talent Availability",
    "strategic_fit":       "Strategic Fit",
}

DIM_DESCRIPTIONS = {
    "market_impact": (
        "Expected business value and demand. Is there strong client demand? "
        "Does it address a major AI adoption challenge? "
        "Will it create competitive advantage for clients?"
    ),
    "effort": (
        "Difficulty and investment to launch and deliver. Consider: technical "
        "complexity, development time, need for new tools or infrastructure, "
        "integration complexity. "
        "SCORE 5 if effort is VERY LOW (quick win). SCORE 1 if VERY HIGH."
    ),
    "scalability": (
        "Can this be delivered repeatedly and efficiently across multiple clients? "
        "Does it follow standardised frameworks, reuse internal accelerators, "
        "require limited customisation per engagement?"
    ),
    "revenue_potential": (
        "Typical deal size and long-term revenue opportunity. Consider: size of "
        "typical engagements, follow-on work potential, long-term managed services "
        "or retainer opportunities."
    ),
    "market_credibility": (
        "Would clients realistically trust an offshore African provider (AmaliTech) "
        "to deliver this? Does AmaliTech have relevant references or adjacent proof "
        "points? Is this typically bought from offshore providers?"
    ),
    "talent_availability": (
        "Does AmaliTech have or can it readily acquire the skills to deliver this "
        "at scale? Consider: availability in Ghana and Rwanda, ease of hiring or "
        "retraining, cost of acquiring expertise."
    ),
    "strategic_fit": (
        "Alignment with AmaliTech's current position. Consider: existing Python/ML, "
        "Azure/cloud, Power BI capabilities; key accounts in manufacturing "
        "(Schaeffler, Knauf), telco (Telekom, 1&1), finance (Serva); Ghana + Rwanda "
        "delivery model; European data privacy and compliance positioning; "
        "vendor-neutral AI stance."
    ),
}

_DIM_GUIDE = "\n".join(
    f"- {DIM_LABELS[d]}: {DIM_DESCRIPTIONS[d]}" for d in DIMENSIONS
)

# ── EXCEL COLOURS ─────────────────────────────────────────────────────────────
NAVY       = "1F3864"
MID_BLUE   = "2E5DA8"
LIGHTER    = "EBF3FA"
WHITE      = "FFFFFF"
HIGH_BG    = "C6EFCE";  HIGH_FG = "276221"
MED_BG     = "FFEB9C";  MED_FG  = "9C6500"
LOW_BG     = "FFC7CE";  LOW_FG  = "9C0006"
GRAY       = "BFBFBF"

_thin   = Side(style="thin", color=GRAY)
BORDER  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ═════════════════════════════════════════════════════════════════════════════
# STATE MANAGEMENT — tracks processed folders so we never double-process
# ═════════════════════════════════════════════════════════════════════════════

def load_state() -> dict:
    """
    Returns dict: { folder_name: { "processed_at": ISO timestamp, "services": int } }
    """
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            log.warning("State file unreadable — starting fresh.")
    return {}


def save_state(state: dict) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    STATE_FILE.write_text(
        json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def clear_state() -> None:
    if STATE_FILE.exists():
        STATE_FILE.unlink()
        log.info("State cleared — all folders will be reprocessed.")


# ═════════════════════════════════════════════════════════════════════════════
# WEIGHTS
# ═════════════════════════════════════════════════════════════════════════════

def parse_weights(raw: str = "") -> dict[str, float]:
    raw = raw.strip() or os.getenv("WEIGHTS", "").strip()
    if raw:
        try:
            vals = [float(x.strip()) for x in raw.split(",")]
            if len(vals) != len(DIMENSIONS):
                log.warning(
                    f"WEIGHTS has {len(vals)} values; expected {len(DIMENSIONS)}. "
                    "Using equal weights."
                )
                raise ValueError
            w = dict(zip(DIMENSIONS, vals))
            log.info(f"Custom weights: {w}")
            return w
        except (ValueError, TypeError):
            pass
    equal = 1.0 / len(DIMENSIONS)
    log.info("Equal weights across all 7 dimensions.")
    return {d: equal for d in DIMENSIONS}


def compute_score(scores: dict, weights: dict) -> float:
    total_w = sum(weights.values()) or 1
    weighted = sum(
        (scores.get(d, {}).get("score", 0) if isinstance(scores.get(d), dict) else 0)
        * weights[d]
        for d in DIMENSIONS
    )
    return round((weighted / (5 * total_w)) * 100, 1)


def priority_tier(score: float) -> str:
    if score >= 70:
        return "High"
    if score >= 45:
        return "Medium"
    return "Low"


def priority_display(score: float) -> str:
    return f"{priority_tier(score)} ({score})"


# ═════════════════════════════════════════════════════════════════════════════
# SITE DISCOVERY
# ═════════════════════════════════════════════════════════════════════════════

def find_site_folders(competitor_filter: str = "") -> list[Path]:
    if not SITES_DIR.exists():
        raise FileNotFoundError(
            f"sites/ not found at {SITES_DIR}. "
            "Run scraper.py or manual_ingest.py first."
        )
    folders = []
    for p in sorted(SITES_DIR.iterdir()):
        if not p.is_dir():
            continue
        has_content = (
            (p / "pages_text.csv").exists() or
            (p / "posts_text.csv").exists()
        )
        if not has_content:
            continue
        if competitor_filter and competitor_filter.lower() not in p.name.lower():
            continue
        folders.append(p)

    if not folders:
        msg = "No site folders with content found"
        if competitor_filter:
            msg += f" matching '{competitor_filter}'"
        raise FileNotFoundError(msg + f" under {SITES_DIR}.")

    return folders


def competitor_name(folder: Path) -> str:
    raw = folder.name.split("_")[0]
    raw = re.sub(r"-(com|ai|world|io|net|org|co)$", "", raw, flags=re.IGNORECASE)
    return raw.replace("-", " ").title()


# ═════════════════════════════════════════════════════════════════════════════
# CONTENT LOADING
# ═════════════════════════════════════════════════════════════════════════════

def load_content(folder: Path, max_pages: int, ocr_min_len: int) -> str:
    chunks: list[str] = []

    for csv_name in ("pages_text.csv", "posts_text.csv"):
        csv_path = folder / csv_name
        if not csv_path.exists():
            continue
        try:
            df = pd.read_csv(csv_path, dtype=str).fillna("").head(max_pages)
            for _, row in df.iterrows():
                url   = row.get("url", row.get("source_image", ""))
                title = row.get("page_title", row.get("source", ""))
                text  = row.get("clean_text", row.get("combined_text",
                        row.get("provided_text", "")))
                if text.strip():
                    chunks.append(f"[PAGE: {title} | {url}]\n{text[:3000]}")
        except Exception as e:
            log.warning(f"Could not read {csv_name} in {folder.name}: {e}")

    ocr_path = folder / "ocr_output.csv"
    if ocr_path.exists():
        try:
            df = pd.read_csv(ocr_path, dtype=str).fillna("")
            for _, row in df.iterrows():
                ocr_text = row.get("extracted_text", "").strip()
                if len(ocr_text) >= ocr_min_len:
                    src = row.get("source_page_url", row.get("source_image", ""))
                    chunks.append(f"[IMAGE TEXT | {src}]\n{ocr_text}")
        except Exception as e:
            log.warning(f"Could not read ocr_output.csv in {folder.name}: {e}")

    return "\n\n---\n\n".join(chunks)


# ═════════════════════════════════════════════════════════════════════════════
# CLAUDE PROMPTS
# ═════════════════════════════════════════════════════════════════════════════

_EXTRACTION_SYSTEM = textwrap.dedent(f"""
    You are a competitive intelligence analyst for AmaliTech — an AI-first
    technology services company operating from Ghana and Rwanda, serving
    European enterprise clients in manufacturing, telecoms, and finance.

    Extract ALL AI-related and AI-adjacent services and products from competitor
    website or social media content. Include:
    - Explicitly AI-branded services (e.g. "AI Readiness Assessment", "LLM Integration")
    - AI-adjacent offerings (ML engineering, data platforms, cloud AI infrastructure,
      automation, analytics, RPA, AI talent placement)
    - Internal AI tools or platforms the competitor has built

    AmaliTech customer maturity levels:
    - AI Explorer: client just starting, no AI in place yet
    - AI Practitioner: client has some AI deployed, actively scaling it
    - AI Champion: client is AI-mature, needs advanced or specialised services
    A service can target a transition e.g. "AI Explorer → AI Practitioner".

    For each service return a JSON object with these exact keys:
    {{
      "name": "short clear service or product name",
      "category": "one of: {' | '.join(SERVICE_CATEGORIES)}",
      "customer_maturity": "one of: {' | '.join(MATURITY_LEVELS)}",
      "description": "1-3 sentences on what this is and who it is for",
      "ai_classification": "core_ai or ai_adjacent",
      "evidence": "short direct quote or specific detail confirming this service",
      "source_url": "page or post URL where found, or empty string"
    }}

    Return ONLY a valid JSON array. No markdown, no preamble, no explanation.
    If nothing relevant is found return: []
""").strip()

_EXTRACTION_USER = "Competitor: {name}\n\nContent:\n{content}"

_SCORING_SYSTEM = textwrap.dedent("""
    You are a strategic analyst for AmaliTech, scoring competitor services against
    AmaliTech's priority matrix to decide which services to build or position against.

    AmaliTech context:
    - Delivery from Ghana + Rwanda; offshore/nearshore engineering + product management
    - Capabilities: Python/ML, Azure/cloud, Power BI, data engineering, QA, DevOps
    - Key accounts: Schaeffler (manufacturing), Deutsche Telekom (telco), Knauf, Serva
    - Target: United Internet (1&1) — telecoms and infrastructure AI
    - Differentiators: European compliance, data sovereignty, vendor-neutral AI, cost advantage

    Score 1–5 per dimension. Ground every justification in AmaliTech's actual context.
    One sentence per justification.

    Dimensions:
    {dim_guide}

    Return ONLY valid JSON, no markdown:
    {{
      "market_impact":        {{"score": <1-5>, "justification": "<one sentence>"}},
      "effort":               {{"score": <1-5>, "justification": "<one sentence>"}},
      "scalability":          {{"score": <1-5>, "justification": "<one sentence>"}},
      "revenue_potential":    {{"score": <1-5>, "justification": "<one sentence>"}},
      "market_credibility":   {{"score": <1-5>, "justification": "<one sentence>"}},
      "talent_availability":  {{"score": <1-5>, "justification": "<one sentence>"}},
      "strategic_fit":        {{"score": <1-5>, "justification": "<one sentence>"}}
    }}
""").strip()

_SCORING_USER = (
    "Competitor: {competitor}\n"
    "Service: {name}\n"
    "Category: {category}\n"
    "Customer Maturity: {maturity}\n"
    "Description: {description}\n"
    "AI Classification: {ai_classification}\n\n"
    "Score this service against AmaliTech's priority matrix."
)


# ═════════════════════════════════════════════════════════════════════════════
# CLAUDE API
# ═════════════════════════════════════════════════════════════════════════════

def _call(client: anthropic.Anthropic, system: str, user: str,
          max_tokens: int = 4096, retries: int = 3) -> str:
    for attempt in range(retries):
        try:
            resp = client.messages.create(
                model=MODEL,
                max_tokens=max_tokens,
                system=system,
                messages=[{"role": "user", "content": user}],
            )
            return resp.content[0].text.strip()
        except anthropic.RateLimitError:
            wait = 30 * (attempt + 1)
            log.warning(f"Rate limit — waiting {wait}s (attempt {attempt+1}/{retries})")
            time.sleep(wait)
        except anthropic.APIError as e:
            log.error(f"API error: {e}")
            if attempt == retries - 1:
                raise
            time.sleep(10)
    raise RuntimeError("All retries exhausted.")


def _parse_json(raw: str, context: str = "") -> dict | list | None:
    clean = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError as e:
        log.error(f"JSON parse error{' for ' + context if context else ''}: {e}")
        log.debug(f"Raw snippet: {clean[:300]}")
        return None


def extract_services(client: anthropic.Anthropic, name: str,
                     content: str) -> list[dict]:
    log.info("  Extracting services...")
    raw = _call(
        client,
        _EXTRACTION_SYSTEM,
        _EXTRACTION_USER.format(name=name, content=content[:60_000]),
        max_tokens=4096,
    )
    result = _parse_json(raw, context=name)
    if not isinstance(result, list):
        log.warning(f"  Unexpected extraction response for {name}.")
        return []
    log.info(f"  {len(result)} service(s) found.")
    return result


def score_service(client: anthropic.Anthropic, competitor: str,
                  service: dict) -> dict:
    raw = _call(
        client,
        _SCORING_SYSTEM.format(dim_guide=_DIM_GUIDE),
        _SCORING_USER.format(
            competitor=competitor,
            name=service.get("name", ""),
            category=service.get("category", ""),
            maturity=service.get("customer_maturity", ""),
            description=service.get("description", ""),
            ai_classification=service.get("ai_classification", ""),
        ),
        max_tokens=1024,
    )
    result = _parse_json(raw, context=service.get("name", ""))
    return result if isinstance(result, dict) else {}


# ═════════════════════════════════════════════════════════════════════════════
# ROW BUILDER
# ═════════════════════════════════════════════════════════════════════════════

def build_rows(competitor: str, services: list[dict],
               all_scores: list[dict], weights: dict) -> list[dict]:
    rows = []
    for svc, scores in zip(services, all_scores):
        score = compute_score(scores, weights)
        row: dict = {
            "competitor":        competitor,
            "service_name":      svc.get("name", ""),
            "category":          svc.get("category", ""),
            "customer_maturity": svc.get("customer_maturity", ""),
            "ai_classification": svc.get("ai_classification", ""),
            "description":       svc.get("description", ""),
            "evidence":          svc.get("evidence", ""),
            "source_url":        svc.get("source_url", ""),
            "priority_score":    score,
            "priority_tier":     priority_tier(score),
            "priority_display":  priority_display(score),
        }
        for dim in DIMENSIONS:
            d = scores.get(dim, {})
            row[f"{dim}_score"] = (
                d.get("score", "") if isinstance(d, dict) else ""
            )
            row[f"{dim}_justification"] = (
                d.get("justification", "") if isinstance(d, dict) else ""
            )
        rows.append(row)

    rows.sort(key=lambda r: r["priority_score"], reverse=True)
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# MASTER CSV — load existing + merge new
# ═════════════════════════════════════════════════════════════════════════════

def load_existing_results() -> pd.DataFrame:
    """Load previously analysed results from the master CSV if it exists."""
    if MASTER_CSV.exists():
        try:
            df = pd.read_csv(MASTER_CSV, dtype=str).fillna("")
            log.info(f"Loaded {len(df)} existing rows from master CSV.")
            return df
        except Exception as e:
            log.warning(f"Could not read existing master CSV: {e}")
    return pd.DataFrame()


def merge_results(existing: pd.DataFrame, new_rows: list[dict]) -> pd.DataFrame:
    """
    Merge new rows into existing results.
    If RERUN_ALL is set, existing rows for reprocessed competitors are dropped first.
    """
    new_df = pd.DataFrame(new_rows) if new_rows else pd.DataFrame()

    if existing.empty:
        return new_df
    if new_df.empty:
        return existing

    # Drop existing rows for competitors being reprocessed
    new_competitors = new_df["competitor"].unique().tolist()
    existing_filtered = existing[
        ~existing["competitor"].isin(new_competitors)
    ]

    merged = pd.concat([existing_filtered, new_df], ignore_index=True)
    return merged.sort_values(
        ["competitor", "priority_score"],
        ascending=[True, False]
    ).reset_index(drop=True)


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═════════════════════════════════════════════════════════════════════════════

def _write_long_list_sheet(ws, rows: list[dict], title: str) -> None:
    """Write Initiative Long List rows to a worksheet."""
    # Title row
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = title
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _fill(NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # Header row
    headers    = ["#", "AI Service Category", "Customer Maturity",
                  "Service", "Description", "Priority"]
    col_widths = [5, 32, 30, 34, 62, 18]

    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font      = _font(bold=True, color=WHITE, size=10)
        cell.fill      = _fill(MID_BLUE)
        cell.alignment = _align("center")
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for row_num, row in enumerate(rows, start=1):
        excel_row = row_num + 2
        tier = row.get("priority_tier", "")
        alt  = row_num % 2 == 0
        bg   = _fill(LIGHTER if alt else WHITE)

        values = [
            row_num,
            row.get("category", ""),
            row.get("customer_maturity", ""),
            row.get("service_name", ""),
            row.get("description", ""),
            row.get("priority_display", ""),
        ]

        for col_i, value in enumerate(values, start=1):
            cell           = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border    = BORDER
            cell.font      = _font(size=9)
            cell.alignment = _align(
                "center" if col_i in (1, 3, 6) else "left",
                wrap=col_i == 5,
            )
            cell.fill = bg

        # Priority cell — colour by tier
        p_cell = ws.cell(row=excel_row, column=6)
        if tier == "High":
            p_cell.fill = _fill(HIGH_BG)
            p_cell.font = _font(bold=True, color=HIGH_FG, size=9)
        elif tier == "Medium":
            p_cell.fill = _fill(MED_BG)
            p_cell.font = _font(bold=True, color=MED_FG, size=9)
        else:
            p_cell.fill = _fill(LOW_BG)
            p_cell.font = _font(bold=True, color=LOW_FG, size=9)

        ws.row_dimensions[excel_row].height = max(
            30, min(80, len(row.get("description", "")) // 4)
        )


def write_initiative_long_list(all_rows: list[dict], path: Path) -> None:
    """
    Write the Initiative Long List xlsx.
    Sheet per competitor (sorted by priority score) + one combined sheet.
    Rows within each competitor are grouped by category, then sorted by score.
    """
    wb = Workbook()
    wb.remove(wb.active)

    by_competitor: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        by_competitor[row["competitor"]].append(row)

    # Combined sheet first
    ws_all = wb.create_sheet(title="All Competitors")
    combined = sorted(all_rows, key=lambda r: (
        r.get("competitor", ""),
        r.get("category", ""),
        -float(r.get("priority_score", 0)),
    ))
    _write_long_list_sheet(ws_all, combined, "Initiative Long List — All Competitors")

    # One sheet per competitor
    for comp_name in sorted(by_competitor.keys()):
        rows = sorted(
            by_competitor[comp_name],
            key=lambda r: (r.get("category", ""), -float(r.get("priority_score", 0))),
        )
        ws = wb.create_sheet(title=comp_name[:28])
        _write_long_list_sheet(ws, rows, f"Initiative Long List — {comp_name}")

    wb.save(path)
    log.info(f"  Initiative Long List: {path.name}  "
             f"({len(all_rows)} services, {len(by_competitor)} competitors)")


def write_detailed_workbook(competitor: str, rows: list[dict], path: Path) -> None:
    """Per-competitor workbook with all dimension scores and justifications."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scored Services"

    total_cols = 7 + len(DIMENSIONS) * 2
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    t = ws["A1"]
    t.value = f"AI Services Analysis — {competitor}"
    t.font  = _font(bold=True, color=WHITE, size=13)
    t.fill  = _fill(NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    base_headers = [
        "Service", "Category", "Customer Maturity", "AI Classification",
        "Description", "Priority Score", "Priority Tier",
    ]
    dim_headers = []
    for d in DIMENSIONS:
        dim_headers += [f"{DIM_LABELS[d]} Score", f"{DIM_LABELS[d]} Justification"]

    all_headers = base_headers + dim_headers
    col_widths  = [30, 28, 24, 16, 50, 14, 12] + [10, 38] * len(DIMENSIONS)

    for col_i, (h, w) in enumerate(zip(all_headers, col_widths), start=1):
        cell           = ws.cell(row=2, column=col_i, value=h)
        cell.font      = _font(bold=True, color=WHITE, size=9)
        cell.fill      = _fill(MID_BLUE)
        cell.alignment = _align("center", wrap=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    for row_num, row in enumerate(rows, start=1):
        excel_row = row_num + 2
        tier = row.get("priority_tier", "")
        alt  = row_num % 2 == 0

        base_values = [
            row.get("service_name", ""),
            row.get("category", ""),
            row.get("customer_maturity", ""),
            row.get("ai_classification", ""),
            row.get("description", ""),
            row.get("priority_score", ""),
            row.get("priority_tier", ""),
        ]
        dim_values = []
        for d in DIMENSIONS:
            dim_values += [
                row.get(f"{d}_score", ""),
                row.get(f"{d}_justification", ""),
            ]

        for col_i, value in enumerate(base_values + dim_values, start=1):
            cell           = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border    = BORDER
            cell.alignment = _align("left", wrap=True)
            cell.font      = _font(size=9)
            cell.fill      = _fill(LIGHTER if alt else WHITE)

        for col_i in (6, 7):
            cell = ws.cell(row=excel_row, column=col_i)
            if tier == "High":
                cell.fill = _fill(HIGH_BG)
                cell.font = _font(bold=True, color=HIGH_FG, size=9)
            elif tier == "Medium":
                cell.fill = _fill(MED_BG)
                cell.font = _font(bold=True, color=MED_FG, size=9)
            else:
                cell.fill = _fill(LOW_BG)
                cell.font = _font(bold=True, color=LOW_FG, size=9)

        ws.row_dimensions[excel_row].height = 40

    wb.save(path)
    log.info(f"  Detailed workbook: {path.name}")


# ═════════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═════════════════════════════════════════════════════════════════════════════

def run(
    competitor_filter: str = "",
    weights_str: str = "",
    max_pages: int = 0,
    ocr_min_len: int = 0,
    dry_run: bool = False,
    rerun_all: bool = False,
) -> pd.DataFrame:
    """
    Incremental pipeline. Returns the full merged master DataFrame.
    Safe to call while scraping is still running.

    rerun_all: if True, ignores state and reprocesses everything.
               Controlled by RERUN_ALL=1 env var or --rerun-all flag.
    """
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key and not dry_run:
        raise EnvironmentError(
            "ANTHROPIC_API_KEY is not set.\n"
            "Run: export ANTHROPIC_API_KEY=your_key_here"
        )

    rerun_all   = rerun_all or os.getenv("RERUN_ALL", "").strip() in ("1", "true", "yes")
    max_pages   = max_pages   or int(os.getenv("MAX_PAGES_PER_SITE", 40))
    ocr_min_len = ocr_min_len or int(os.getenv("OCR_MIN_LENGTH", 30))
    weights     = parse_weights(weights_str)

    OUTPUT_DIR.mkdir(exist_ok=True)

    # ── Load state ──
    if rerun_all:
        clear_state()
    state = load_state()

    # ── Discover all site folders ──
    all_folders = find_site_folders(competitor_filter)

    # ── Separate pending vs already done ──
    pending   = [f for f in all_folders if f.name not in state]
    completed = [f for f in all_folders if f.name in state]

    log.info(f"\nSite folders found : {len(all_folders)}")
    log.info(f"  Already analysed : {len(completed)}")
    log.info(f"  Pending          : {len(pending)}")

    if completed:
        log.info("  Completed folders:")
        for f in completed:
            info = state[f.name]
            log.info(f"    ✓ {competitor_name(f):<25} "
                     f"({info.get('services', '?')} services, "
                     f"processed {info.get('processed_at', '?')})")

    if pending:
        log.info("  Pending folders:")
        for f in pending:
            log.info(f"    … {competitor_name(f)}")

    if dry_run:
        log.info("\nDRY RUN — no API calls made.")
        return pd.DataFrame()

    if not pending:
        log.info("\nNo new folders to process.")
        log.info("Rebuilding outputs from existing results...")
        existing = load_existing_results()
        if existing.empty:
            log.warning("No existing results found either. Nothing to output.")
            return pd.DataFrame()
        _write_outputs(existing, state)
        return existing

    # ── Process pending folders ──
    client   = anthropic.Anthropic(api_key=api_key)
    new_rows: list[dict] = []

    for folder in pending:
        name = competitor_name(folder)
        log.info(f"\n{'─'*60}")
        log.info(f"  {name}  ({folder.name})")
        log.info(f"{'─'*60}")

        content = load_content(folder, max_pages, ocr_min_len)
        if not content.strip():
            log.warning("  No content found — skipping.")
            # Still mark as processed so we don't retry empty folders endlessly
            state[folder.name] = {
                "processed_at": datetime.now().isoformat(timespec="seconds"),
                "services": 0,
                "competitor": name,
                "skipped": True,
            }
            save_state(state)
            continue

        log.info(f"  Content: ~{len(content):,} chars")

        services = extract_services(client, name, content)
        if not services:
            log.warning("  No services extracted — skipping.")
            state[folder.name] = {
                "processed_at": datetime.now().isoformat(timespec="seconds"),
                "services": 0,
                "competitor": name,
                "skipped": True,
            }
            save_state(state)
            continue

        scored: list[dict] = []
        for i, svc in enumerate(services, 1):
            log.info(f"  Scoring [{i}/{len(services)}]: {svc.get('name', '?')}")
            scores = score_service(client, name, svc)
            scored.append(scores)
            time.sleep(0.5)

        rows = build_rows(name, services, scored, weights)
        new_rows.extend(rows)

        # Per-competitor detailed xlsx
        safe       = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
        detail_path = OUTPUT_DIR / f"{safe}_services_scored.xlsx"
        write_detailed_workbook(name, rows, detail_path)

        # Update state
        state[folder.name] = {
            "processed_at": datetime.now().isoformat(timespec="seconds"),
            "services":     len(rows),
            "competitor":   name,
            "skipped":      False,
        }
        save_state(state)
        log.info(f"  State saved — {name} marked as done.")

    # ── Merge new + existing results ──
    existing = load_existing_results()
    master   = merge_results(existing, new_rows)

    if master.empty:
        log.warning("No results to write.")
        return master

    # ── Write outputs ──
    _write_outputs(master, state)
    return master


def _write_outputs(master: pd.DataFrame, state: dict) -> None:
    """Write master CSV and Initiative Long List xlsx from the full merged DataFrame."""

    # Master CSV
    master.to_csv(MASTER_CSV, index=False, encoding="utf-8")
    log.info(f"  Master CSV updated: {MASTER_CSV.name}  ({len(master)} rows)")

    # Initiative Long List xlsx — always named with today's date
    ts      = datetime.now().strftime("%Y%m%d")
    ll_path = OUTPUT_DIR / f"{ts}_initiative_long_list.xlsx"
    write_initiative_long_list(master.to_dict("records"), ll_path)

    # Summary
    competitors_done = sum(
        1 for v in state.values() if not v.get("skipped", False)
    )
    total_sites = len(state)

    log.info(f"\n{'='*60}")
    log.info("OUTPUT SUMMARY")
    log.info(f"  Competitors processed : {competitors_done} / {total_sites} sites tracked")
    log.info(f"  Total services        : {len(master)}")
    for tier in ["High", "Medium", "Low"]:
        n = (master["priority_tier"] == tier).sum()
        log.info(f"  {tier:<8} priority    : {n}")
    log.info(f"  Outputs in            : {OUTPUT_DIR}/")
    log.info(f"    {ll_path.name}")
    log.info(f"    {MASTER_CSV.name}")
    log.info(f"    processed_folders.json  (state tracker)")
    log.info(f"{'='*60}\n")

    # Print pending sites still not scraped/analysed
    all_site_names = {f.name for f in SITES_DIR.iterdir() if f.is_dir()} \
        if SITES_DIR.exists() else set()
    not_yet = all_site_names - set(state.keys())
    if not_yet:
        log.info(f"  Folders in sites/ not yet analysed ({len(not_yet)}):")
        for n in sorted(not_yet):
            log.info(f"    - {n}")


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="AmaliTech Competitor AI Services — Incremental Analysis"
    )
    parser.add_argument(
        "--competitor", "-c", default="",
        help="Filter to a single competitor by name fragment (e.g. 'andela')",
    )
    parser.add_argument(
        "--weights", "-w", default="",
        help=(
            "Comma-separated weights for 7 dimensions: "
            "market_impact,effort,scalability,revenue_potential,"
            "market_credibility,talent_availability,strategic_fit"
        ),
    )
    parser.add_argument(
        "--max-pages", "-p", type=int, default=0,
        help="Max pages to load per competitor (default: 40)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Show pending/completed folders without making API calls",
    )
    parser.add_argument(
        "--rerun-all", action="store_true",
        help="Ignore state and reprocess all folders from scratch "
             "(also controlled by RERUN_ALL=1 env var)",
    )
    args = parser.parse_args()

    run(
        competitor_filter=args.competitor,
        weights_str=args.weights,
        max_pages=args.max_pages,
        dry_run=args.dry_run,
        rerun_all=args.rerun_all,
    )


if __name__ == "__main__":
    main()
