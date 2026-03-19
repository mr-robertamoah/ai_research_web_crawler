"""
AmaliTech Competitor AI Services — Analysis Script (Groq edition)
=================================================================
Drop-in replacement for analyse_ollama.py using Groq's OpenAI-compatible
API instead of local Ollama. Everything else — incremental state tracking,
scoring, Excel output — is identical.

HOW TO RUN:
  python analyse_groq.py                        # process only NEW site folders
  python analyse_groq.py --competitor andela    # single competitor
  python analyse_groq.py --max-pages 10         # faster / lower token use
  python analyse_groq.py --dry-run              # show state, no model calls
  python analyse_groq.py --rerun-all            # reprocess everything

ENVIRONMENT VARIABLES:
  GROQ_API_KEY          Your Groq API key (required)
  GROQ_MODEL            Model to use (default: qwen/qwen3-32b)
  RERUN_ALL             set to 1 to reprocess everything
  WEIGHTS               comma-separated weights for 7 dimensions
  MAX_PAGES_PER_SITE    max pages to feed per competitor (default: 40)
  OCR_MIN_LENGTH        min chars for OCR row to be included (default: 30)
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import time
import textwrap
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("analyse_ollama")

# ── PATHS ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(os.getenv("APP_DIR", Path(__file__).parent.resolve()))
SITES_DIR  = SCRIPT_DIR / "sites"
OUTPUT_DIR = SCRIPT_DIR / "output"
STATE_FILE      = OUTPUT_DIR / "processed_folders_groq.json"
MASTER_CSV      = OUTPUT_DIR / "all_competitors_priority_groq.csv"
HYPOTHESIS_FILE = OUTPUT_DIR / "hypothesis_tracker_groq.json"

# ── GROQ ──────────────────────────────────────────────────────────────────────
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
GROQ_MODEL   = os.getenv("GROQ_MODEL", "qwen/qwen3-32b")
GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"

# ── CATEGORIES / MATURITY / DIMENSIONS ───────────────────────────────────────
SERVICE_CATEGORIES = [
    "AI Advisory & Readiness",
    "AI Engineering & Automation",
    "AI Platforms & Agents",
    "AI-powered Solutions & New Revenue Models",
    "Talent & Staffing",
    "Other",
]

MATURITY_LEVELS = [
    "AI Explorer",
    "AI Practitioner",
    "AI Champion",
    "AI Explorer → AI Practitioner",
    "AI Practitioner → AI Champion",
]

DIMENSIONS = [
    "market_impact",
    "effort",
    "scalability",
    "revenue_potential",
    "market_credibility",
    "talent_availability",
    "strategic_fit",
]

DIM_LABELS = {
    "market_impact":       "Market Impact",
    "effort":              "Effort (inverse: 5=low effort)",
    "scalability":         "Scalability",
    "revenue_potential":   "Revenue Potential",
    "market_credibility":  "Market Credibility",
    "talent_availability": "Talent Availability",
    "strategic_fit":       "Strategic Fit",
}

DIM_DESCRIPTIONS = {
    "market_impact":       "Expected business value and demand. Strong client demand? Addresses major AI adoption challenge?",
    "effort":              "Difficulty to launch. SCORE 5 = VERY LOW effort (quick win). SCORE 1 = VERY HIGH effort.",
    "scalability":         "Can this be delivered repeatedly across multiple clients with standardised frameworks?",
    "revenue_potential":   "Typical deal size, follow-on work, recurring managed services opportunity.",
    "market_credibility":  "Would clients trust an offshore African provider (AmaliTech) to deliver this?",
    "talent_availability": "Skills available in Ghana/Rwanda, ease of hiring/training, cost of expertise.",
    "strategic_fit":       "Alignment with existing accounts (Schaeffler, Telekom, Knauf), capabilities, European compliance.",
}

_DIM_GUIDE = "\n".join(
    f"- {DIM_LABELS[d]}: {DIM_DESCRIPTIONS[d]}" for d in DIMENSIONS
)

# ── EXCEL COLOURS ─────────────────────────────────────────────────────────────
NAVY     = "1F3864"; MID_BLUE = "2E5DA8"; LIGHTER = "EBF3FA"; WHITE = "FFFFFF"
HIGH_BG  = "C6EFCE"; HIGH_FG = "276221"
MED_BG   = "FFEB9C"; MED_FG  = "9C6500"
LOW_BG   = "FFC7CE"; LOW_FG  = "9C0006"
GRAY     = "BFBFBF"

_thin  = Side(style="thin", color=GRAY)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ═════════════════════════════════════════════════════════════════════════════
# STATE
# ═════════════════════════════════════════════════════════════════════════════

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            log.warning("State file unreadable — starting fresh.")
    return {}

def save_state(state: dict) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")

def clear_state() -> None:
    if STATE_FILE.exists():
        STATE_FILE.unlink()
        log.info("State cleared — all folders will be reprocessed.")


# ═════════════════════════════════════════════════════════════════════════════
# WEIGHTS & SCORING
# ═════════════════════════════════════════════════════════════════════════════

def parse_weights(raw: str = "") -> dict[str, float]:
    raw = raw.strip() or os.getenv("WEIGHTS", "").strip()
    if raw:
        try:
            vals = [float(x.strip()) for x in raw.split(",")]
            if len(vals) == len(DIMENSIONS):
                return dict(zip(DIMENSIONS, vals))
        except (ValueError, TypeError):
            pass
    equal = 1.0 / len(DIMENSIONS)
    return {d: equal for d in DIMENSIONS}

def compute_score(scores: dict, weights: dict) -> float:
    total_w = sum(weights.values()) or 1
    weighted = sum(
        (scores.get(d, {}).get("score", 0) if isinstance(scores.get(d), dict) else 0)
        * weights[d]
        for d in DIMENSIONS
    )
    return round((weighted / (5 * total_w)) * 100, 1)

def priority_tier(score: float) -> str:
    if score >= 70: return "High"
    if score >= 45: return "Medium"
    return "Low"

def priority_display(score: float) -> str:
    return f"{priority_tier(score)} ({score})"


# ═════════════════════════════════════════════════════════════════════════════
# SITE DISCOVERY
# ═════════════════════════════════════════════════════════════════════════════

def find_site_folders(competitor_filter: str = "") -> list[Path]:
    if not SITES_DIR.exists():
        raise FileNotFoundError(f"sites/ not found at {SITES_DIR}.")
    folders = []
    for p in sorted(SITES_DIR.iterdir()):
        if not p.is_dir():
            continue
        has_content = (p / "pages_text.csv").exists() or (p / "posts_text.csv").exists()
        if not has_content:
            continue
        if competitor_filter and competitor_filter.lower() not in p.name.lower():
            continue
        folders.append(p)
    if not folders:
        raise FileNotFoundError(f"No site folders with content found under {SITES_DIR}.")
    return folders

def competitor_name(folder: Path) -> str:
    raw = folder.name.split("_")[0]
    raw = re.sub(r"-(com|ai|world|io|net|org|co)$", "", raw, flags=re.IGNORECASE)
    return raw.replace("-", " ").title()


# ═════════════════════════════════════════════════════════════════════════════
# CONTENT LOADING
# ═════════════════════════════════════════════════════════════════════════════

def load_content(folder: Path, max_pages: int, ocr_min_len: int) -> str:
    chunks: list[str] = []
    for csv_name in ("pages_text.csv", "posts_text.csv"):
        csv_path = folder / csv_name
        if not csv_path.exists():
            continue
        try:
            df = pd.read_csv(csv_path, dtype=str).fillna("").head(max_pages)
            for _, row in df.iterrows():
                url   = row.get("url", row.get("source_image", ""))
                title = row.get("page_title", row.get("source", ""))
                text  = row.get("clean_text", row.get("combined_text", row.get("provided_text", "")))
                if text.strip():
                    chunks.append(f"[PAGE: {title} | {url}]\n{text[:3000]}")
        except Exception as e:
            log.warning(f"Could not read {csv_name} in {folder.name}: {e}")

    ocr_path = folder / "ocr_output.csv"
    if ocr_path.exists():
        try:
            df = pd.read_csv(ocr_path, dtype=str).fillna("")
            for _, row in df.iterrows():
                ocr_text = row.get("extracted_text", "").strip()
                if len(ocr_text) >= ocr_min_len:
                    src = row.get("source_page_url", row.get("source_image", ""))
                    chunks.append(f"[IMAGE TEXT | {src}]\n{ocr_text}")
        except Exception as e:
            log.warning(f"Could not read ocr_output.csv in {folder.name}: {e}")

    return "\n\n---\n\n".join(chunks)

# ═════════════════════════════════════════════════════════════════════════════
# OLLAMA API
# ═════════════════════════════════════════════════════════════════════════════

def _call_ollama(system: str, user: str, retries: int = 3) -> str:
    """Call Groq OpenAI-compatible API and return the assistant message text."""
    if not GROQ_API_KEY:
        raise RuntimeError("GROQ_API_KEY environment variable is not set.")
    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": GROQ_MODEL,
        "temperature": 0.1,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
    }
    for attempt in range(retries):
        try:
            resp = requests.post(GROQ_API_URL, json=payload, headers=headers, timeout=120)
            resp.raise_for_status()
            return resp.json()["choices"][0]["message"]["content"].strip()
        except requests.exceptions.HTTPError as e:
            log.warning(f"Groq API error (attempt {attempt+1}/{retries}): {e} — {resp.text[:200]}")
        except Exception as e:
            log.warning(f"Groq call failed (attempt {attempt+1}/{retries}): {e}")
        if attempt == retries - 1:
            raise RuntimeError("All Groq retries exhausted.")
        time.sleep(15 * (attempt + 1))
    raise RuntimeError("All Groq retries exhausted.")


def _parse_json(raw: str, context: str = "") -> dict | list | None:
    # Strip markdown code fences if present
    clean = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
    clean = re.sub(r"\s*```$", "", clean.strip(), flags=re.MULTILINE)
    # Strip <think>...</think> blocks that qwen3 sometimes emits
    clean = re.sub(r"<think>.*?</think>", "", clean, flags=re.DOTALL).strip()
    # Normalise pipe-separated enum values inside JSON strings (e.g. "A | B" → "A")
    clean = re.sub(r'"([^"]*)\s*\|\s*([^"]*)"', lambda m: f'"{m.group(1).strip()}"', clean)
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        # Try to find the first JSON array or object in the response
        for pattern in (r"(\[.*\])", r"(\{.*\})"):
            m = re.search(pattern, clean, re.DOTALL)
            if m:
                try:
                    return json.loads(m.group(1))
                except json.JSONDecodeError:
                    pass
        log.error(f"JSON parse error{' for ' + context if context else ''}. Snippet: {clean[:500]}")
        # Last resort: try to salvage a truncated JSON array by closing it
        m = re.search(r"(\[.*)", clean, re.DOTALL)
        if m:
            partial = m.group(1)
            # Close any open object and the array
            open_braces = partial.count("{") - partial.count("}")
            partial += "}" * open_braces + "]"
            try:
                return json.loads(partial)
            except json.JSONDecodeError:
                pass
        return None


# ── PROMPTS ───────────────────────────────────────────────────────────────────

HYPOTHESES = [
    "Competitors are charging an AI premium of 15–30% over baseline managed services rates.",
    "The fastest-growing competitors are pivoting from time-and-materials to outcome/value-based pricing.",
    "AI capability is being built primarily through hyperscaler partnerships (OpenAI, Anthropic, AWS, Google, Azure) rather than internal R&D.",
    "European enterprise buyers are prioritising data sovereignty and compliance-safe AI — creating an opening for AmaliTech's delivery model.",
    "Competitors are concentrating AI investment in 2–3 verticals rather than spreading across all sectors.",
]

_EXTRACTION_SYSTEM = textwrap.dedent(f"""
    You are a competitive intelligence analyst for AmaliTech — an AI-first
    technology services company from Ghana and Rwanda serving European enterprise
    clients in manufacturing, telecoms, and finance.

    Extract ALL AI-related and AI-adjacent services from the competitor content.
    Include: AI-branded services, ML engineering, data platforms, cloud AI,
    automation, analytics, RPA, AI talent placement, internal AI tools.

    Customer maturity levels:
    - AI Explorer: client just starting, no AI in place
    - AI Practitioner: client has some AI, actively scaling
    - AI Champion: AI-mature, needs advanced services
    - AI Explorer → AI Practitioner (transition)
    - AI Practitioner → AI Champion (transition)

    Return ONLY a valid JSON array. No markdown, no explanation, no preamble.
    Each element must have exactly these keys:
    {{
      "name": "short service name",
      "category": "one of: {' | '.join(SERVICE_CATEGORIES)}",
      "customer_maturity": "one of: {' | '.join(MATURITY_LEVELS)}",
      "description": "1-3 sentences on what this is and who it is for",
      "plain_english_summary": "1 sentence explaining what this does in simple, jargon-free language a non-technical executive would understand",
      "ai_classification": "core_ai or ai_adjacent",
      "pricing_signals": "any pricing model hints, contract type, or AI premium mentions found — or empty string",
      "client_wins": "named clients or industries mentioned as customers — or empty string",
      "tech_stack": "LLM providers, cloud platforms, or proprietary tools mentioned — or empty string",
      "evidence": "short quote or detail confirming this service",
      "source_url": "page URL where found, or empty string"
    }}

    If nothing relevant is found return: []
""").strip()

_HYPOTHESIS_SYSTEM = (
    "You are a competitive intelligence analyst for AmaliTech — a social enterprise IT services company "
    "delivering from Ghana and Rwanda to European enterprise clients. "
    "AmaliTech is ISO 27001/TISAX certified, an AWS Advanced Partner, and is positioning as AI-first. "
    "Given competitor content, assess evidence for/against each hypothesis. "
    "Return ONLY valid JSON, no markdown:\n"
    '{"h1":{"evidence_for":"<quote or finding, or empty>","evidence_against":"<quote or finding, or empty>","verdict":"Confirmed|Refuted|Insufficient data"},'
    '"h2":{"evidence_for":"","evidence_against":"","verdict":"Confirmed|Refuted|Insufficient data"},'
    '"h3":{"evidence_for":"","evidence_against":"","verdict":"Confirmed|Refuted|Insufficient data"},'
    '"h4":{"evidence_for":"","evidence_against":"","verdict":"Confirmed|Refuted|Insufficient data"},'
    '"h5":{"evidence_for":"","evidence_against":"","verdict":"Confirmed|Refuted|Insufficient data"}}'
)

_SCORING_SYSTEM = (
    "You are scoring competitor AI services for AmaliTech's strategic priority matrix.\n\n"
    "AmaliTech context:\n"
    "- Social enterprise (gGmbH), ~400 staff, delivery from Ghana + Rwanda, HQ Cologne Germany\n"
    "- Engagement models: dedicated teams, project-based, hybrid — embedded in client teams\n"
    "- Certifications: ISO 9001, ISO/IEC 27001, TISAX (highest score) — strong European data sovereignty positioning\n"
    "- AWS Advanced Partner, 100+ AWS-certified engineers, 30+ cloud projects\n"
    "- Tech stack: AWS (primary), Python, JavaScript/React, Node.js, PHP, Azure, Power BI, SAP BTP, Salesforce, iOS/Android\n"
    "- Training Academy: 75 trainees/quarter in frontend, backend, DevOps, data engineering, QA, mobile, AWS\n"
    "- AmaliAI: internal AI product wrapping popular LLMs (OpenAI, AWS Bedrock) — not yet client-facing\n\n"
    "CURRENT CAPABILITIES (effort score 4-5 — deliverable now):\n"
    "  • Software development (frontend, backend, mobile, full-stack) with embedded AI features\n"
    "  • End-to-end data pipelines, data engineering, BI and analytics (Power BI)\n"
    "  • LLM/AI integration by wrapping AWS Bedrock, Azure OpenAI, or open-source APIs into applications\n"
    "  • RAG pipelines and simple AI agents via LangChain/LlamaIndex\n"
    "  • Cloud transformation and DevOps on AWS/Azure\n"
    "  • QA, cybersecurity assessment, SAP BTP integration, IT support\n"
    "  • AI advisory and readiness assessments (a few senior consultants available)\n\n"
    "ROADMAP 6-12 MONTHS (effort score 3 — requires upskilling, learnable):\n"
    "  • MLOps/LLMOps pipelines, model observability and monitoring\n"
    "  • Agentic AI systems with complex multi-agent orchestration\n"
    "  • AI-powered product management and outcome-based delivery\n\n"
    "ROADMAP 1-2 YEARS (effort score 1-2 — significant new capability needed):\n"
    "  • Custom model training or fine-tuning from scratch\n"
    "  • Proprietary AI platform development\n"
    "  • Deep vertical AI (industrial computer vision, autonomous systems, biomedical AI)\n"
    "  • Large-scale RLHF or human feedback data operations\n\n"
    "TARGET INDUSTRIES AND CLIENT MATURITY:\n"
    "  • Manufacturing (Schaeffler-type, AI Explorer → Practitioner): greenfield AI; will pay for clear ROI use cases "
    "like predictive maintenance dashboards, quality control analytics, supply chain visibility — built on existing data pipelines\n"
    "  • Telecoms (DTIT = existing software dev + data pipelines, AI Explorer → Practitioner; United Internet = target new logo): "
    "will pay for network analytics, AI-enhanced DevOps, customer data platforms with clear cost/efficiency ROI\n"
    "  • E-commerce (aspirational, no current clients): will pay for demand forecasting, personalisation, fraud detection\n"
    "  • Score strategic_fit HIGHEST for services that: (a) fit manufacturing/telecoms/e-commerce, "
    "(b) have a clear budgeted problem with demonstrable ROI (cost reduction or revenue uplift), "
    "AND (c) build naturally on AmaliTech's existing software dev + data pipeline work\n"
    "  • Score strategic_fit LOWER for services requiring capabilities AmaliTech doesn't have or industries outside these three\n\n"
    "Score 1–5 per dimension. Effort is INVERSE (5=deliverable now, 1=2+ year roadmap).\n\n"
    "Return ONLY valid JSON:\n"
    '{"market_impact":{"score":<1-5>,"justification":"<one sentence — reference which industry and client maturity this serves>"},'
    '"effort":{"score":<1-5>,"justification":"<one sentence — explicitly state: deliverable now / 6-12mo roadmap / 1-2yr roadmap>"},'
    '"scalability":{"score":<1-5>,"justification":"<one sentence>"},'
    '"revenue_potential":{"score":<1-5>,"justification":"<one sentence — reference deal size or ROI story>"},'
    '"market_credibility":{"score":<1-5>,"justification":"<one sentence — reference ISO/TISAX/AWS and whether Explorer/Practitioner clients would trust AmaliTech>"},'
    '"talent_availability":{"score":<1-5>,"justification":"<one sentence — reference Ghana/Rwanda talent and training academy>"},'
    '"strategic_fit":{"score":<1-5>,"justification":"<one sentence — reference specific accounts, target industries, and ROI fit>"}}'
)


def data_confidence(content: str) -> str:
    length = len(content)
    if length >= 30_000: return "High"
    if length >= 8_000:  return "Medium"
    return "Low"


def _sanitise_services(services: list[dict]) -> list[dict]:
    """Clamp enum fields to valid values so bad model output doesn't crash downstream."""
    for svc in services:
        raw_maturity = svc.get("customer_maturity", "")
        if raw_maturity not in MATURITY_LEVELS:
            match = next((m for m in MATURITY_LEVELS if m.lower() in raw_maturity.lower()), MATURITY_LEVELS[0])
            svc["customer_maturity"] = match
        if svc.get("category", "") not in SERVICE_CATEGORIES:
            svc["category"] = "Other"
    return services


_AI_KEYWORDS = re.compile(
    r"\b(ai|ml|machine learning|deep learning|llm|generative|automation|analytics|"
    r"data science|nlp|computer vision|predictive|intelligent|chatbot|copilot|"
    r"neural|model|inference|embedding|vector|rag|fine.tun)\b",
    re.IGNORECASE,
)

def _smart_excerpt(content: str, limit: int = 6_000) -> str:
    """Return up to `limit` chars, prioritising page blocks that mention AI keywords."""
    blocks = content.split("\n\n---\n\n")
    # Score each block by AI keyword density
    scored = sorted(blocks, key=lambda b: len(_AI_KEYWORDS.findall(b)), reverse=True)
    result, used = [], 0
    for block in scored:
        take = block[:limit - used]
        result.append(take)
        used += len(take)
        if used >= limit:
            break
    return "\n\n---\n\n".join(result)


def extract_services(name: str, content: str) -> list[dict]:
    log.info("  Extracting services...")
    raw = _call_ollama(
        _EXTRACTION_SYSTEM,
        f"Competitor: {name}\n\nContent:\n{_smart_excerpt(content)}",
    )
    result = _parse_json(raw, context=name)
    if not isinstance(result, list):
        log.warning(f"  Unexpected extraction response for {name}.")
        return []
    result = _sanitise_services(result)
    log.info(f"  {len(result)} service(s) found.")
    return result


def assess_hypotheses(name: str, content: str) -> dict:
    log.info("  Assessing hypotheses...")
    hyp_context = "\n".join(f"H{i+1}: {h}" for i, h in enumerate(HYPOTHESES))
    raw = _call_ollama(
        _HYPOTHESIS_SYSTEM,
        f"Competitor: {name}\nHypotheses:\n{hyp_context}\n\nContent:\n{_smart_excerpt(content)}",
    )
    result = _parse_json(raw, context=f"{name} hypotheses")
    return result if isinstance(result, dict) else {}


def score_service(competitor: str, service: dict) -> dict:
    user = (
        f"Competitor: {competitor}\n"
        f"Service: {service.get('name', '')}\n"
        f"Category: {service.get('category', '')}\n"
        f"Customer Maturity: {service.get('customer_maturity', '')}\n"
        f"Description: {service.get('description', '')}\n"
        f"AI Classification: {service.get('ai_classification', '')}\n\n"
        "Score this service against AmaliTech's priority matrix."
    )
    raw    = _call_ollama(_SCORING_SYSTEM, user)
    result = _parse_json(raw, context=service.get("name", ""))
    return result if isinstance(result, dict) else {}

# ═════════════════════════════════════════════════════════════════════════════
# ROW BUILDER
# ═════════════════════════════════════════════════════════════════════════════

def build_rows(competitor: str, services: list[dict],
               all_scores: list[dict], weights: dict,
               confidence: str = "Medium") -> list[dict]:
    rows = []
    for svc, scores in zip(services, all_scores):
        score = compute_score(scores, weights)
        row: dict = {
            "competitor":              competitor,
            "service_name":            svc.get("name", ""),
            "category":                svc.get("category", ""),
            "customer_maturity":       svc.get("customer_maturity", ""),
            "ai_classification":       svc.get("ai_classification", ""),
            "description":             svc.get("description", ""),
            "plain_english_summary":   svc.get("plain_english_summary", ""),
            "pricing_signals":         svc.get("pricing_signals", ""),
            "client_wins":             svc.get("client_wins", ""),
            "tech_stack":              svc.get("tech_stack", ""),
            "data_confidence":         confidence,
            "evidence":                svc.get("evidence", ""),
            "source_url":              svc.get("source_url", ""),
            "priority_score":    score,
            "priority_tier":     priority_tier(score),
            "priority_display":  priority_display(score),
        }
        for dim in DIMENSIONS:
            d = scores.get(dim, {})
            row[f"{dim}_score"]         = d.get("score", "")         if isinstance(d, dict) else ""
            row[f"{dim}_justification"] = d.get("justification", "") if isinstance(d, dict) else ""
        rows.append(row)
    rows.sort(key=lambda r: r["priority_score"], reverse=True)
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# MASTER CSV
# ═════════════════════════════════════════════════════════════════════════════

def load_hypothesis_data() -> dict:
    if HYPOTHESIS_FILE.exists():
        try:
            return json.loads(HYPOTHESIS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_hypothesis_data(data: dict) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    HYPOTHESIS_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def load_existing_results() -> pd.DataFrame:
    if MASTER_CSV.exists():
        try:
            df = pd.read_csv(MASTER_CSV, dtype=str).fillna("")
            log.info(f"Loaded {len(df)} existing rows from master CSV.")
            return df
        except Exception as e:
            log.warning(f"Could not read existing master CSV: {e}")
    return pd.DataFrame()

def merge_results(existing: pd.DataFrame, new_rows: list[dict]) -> pd.DataFrame:
    new_df = pd.DataFrame(new_rows) if new_rows else pd.DataFrame()
    if existing.empty:
        return new_df
    if new_df.empty:
        return existing
    new_competitors = new_df["competitor"].unique().tolist()
    existing_filtered = existing[~existing["competitor"].isin(new_competitors)]
    merged = pd.concat([existing_filtered, new_df], ignore_index=True)
    return merged.sort_values(["competitor", "priority_score"],
                              ascending=[True, False]).reset_index(drop=True)
    if MASTER_CSV.exists():
        try:
            df = pd.read_csv(MASTER_CSV, dtype=str).fillna("")
            log.info(f"Loaded {len(df)} existing rows from master CSV.")
            return df
        except Exception as e:
            log.warning(f"Could not read existing master CSV: {e}")
    return pd.DataFrame()

def merge_results(existing: pd.DataFrame, new_rows: list[dict]) -> pd.DataFrame:
    new_df = pd.DataFrame(new_rows) if new_rows else pd.DataFrame()
    if existing.empty:
        return new_df
    if new_df.empty:
        return existing
    new_competitors = new_df["competitor"].unique().tolist()
    existing_filtered = existing[~existing["competitor"].isin(new_competitors)]
    merged = pd.concat([existing_filtered, new_df], ignore_index=True)
    return merged.sort_values(["competitor", "priority_score"],
                              ascending=[True, False]).reset_index(drop=True)


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═════════════════════════════════════════════════════════════════════════════

def _write_long_list_sheet(ws, rows: list[dict], title: str) -> None:
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = title
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _fill(NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers    = ["#", "AI Service Category", "Customer Maturity", "Service", "Description", "Plain English Summary", "Priority"]
    col_widths = [5, 32, 30, 34, 62, 62, 18]
    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _align("center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    for row_num, row in enumerate(rows, start=1):
        excel_row = row_num + 2
        tier = row.get("priority_tier", "")
        bg   = _fill(LIGHTER if row_num % 2 == 0 else WHITE)
        values = [row_num, row.get("category",""), row.get("customer_maturity",""),
                  row.get("service_name",""), row.get("description",""),
                  row.get("plain_english_summary",""), row.get("priority_display","")]
        for col_i, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border = BORDER
            cell.font = _font(size=9)
            cell.alignment = _align("center" if col_i in (1,3,7) else "left", wrap=col_i in (5,6))
            cell.fill = bg
        p = ws.cell(row=excel_row, column=7)
        if tier == "High":
            p.fill = _fill(HIGH_BG); p.font = _font(bold=True, color=HIGH_FG, size=9)
        elif tier == "Medium":
            p.fill = _fill(MED_BG);  p.font = _font(bold=True, color=MED_FG,  size=9)
        else:
            p.fill = _fill(LOW_BG);  p.font = _font(bold=True, color=LOW_FG,  size=9)
        ws.row_dimensions[excel_row].height = max(30, min(80, len(row.get("description","")) // 4))


def _write_hypothesis_sheet(ws, hypothesis_data: dict) -> None:
    """Write hypothesis tracker sheet."""
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Hypothesis Tracker"
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _fill(NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Hypothesis", "Competitor", "Evidence For", "Evidence Against", "Verdict"]
    widths  = [5, 60, 22, 50, 50, 20]
    for col_i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _align("center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    excel_row = 3
    for h_idx, hyp_text in enumerate(HYPOTHESES, start=1):
        h_key = f"h{h_idx}"
        comp_entries = [(comp, data[h_key]) for comp, data in hypothesis_data.items() if h_key in data]
        if not comp_entries:
            comp_entries = [("—", {"evidence_for": "", "evidence_against": "", "verdict": "Insufficient data"})]
        for comp, entry in comp_entries:
            verdict = entry.get("verdict", "Insufficient data")
            bg = _fill(HIGH_BG if verdict == "Confirmed" else LOW_BG if verdict == "Refuted" else MED_BG)
            values = [h_idx, hyp_text, comp, entry.get("evidence_for",""), entry.get("evidence_against",""), verdict]
            for col_i, val in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_i, value=val)
                cell.border = BORDER
                cell.font = _font(size=9)
                cell.alignment = _align("left", wrap=True)
            ws.cell(row=excel_row, column=6).fill = bg
            ws.cell(row=excel_row, column=6).font = _font(bold=True, size=9)
            ws.row_dimensions[excel_row].height = 40
            excel_row += 1


def _write_comparison_matrix_sheet(ws, all_rows: list[dict]) -> None:
    """Write comparison matrix — one row per competitor, avg scores per dimension."""
    ws.merge_cells(f"A1:{get_column_letter(2 + len(DIMENSIONS))}1")
    c = ws["A1"]
    c.value = "Competitor Comparison Matrix"
    c.font = _font(bold=True, color=WHITE, size=13)
    c.fill = _fill(NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers = ["Competitor", "Avg Priority Score"] + [DIM_LABELS[d] for d in DIMENSIONS] + ["Data Confidence", "Services Count"]
    widths  = [28, 18] + [16]*len(DIMENSIONS) + [16, 14]
    for col_i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font = _font(bold=True, color=WHITE, size=9)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _align("center", wrap=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    # Aggregate by competitor
    by_comp: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        by_comp[row["competitor"]].append(row)

    for row_num, comp in enumerate(sorted(by_comp.keys()), start=1):
        rows = by_comp[comp]
        excel_row = row_num + 2
        avg_score = round(sum(float(r.get("priority_score", 0)) for r in rows) / len(rows), 1)
        dim_avgs  = []
        for d in DIMENSIONS:
            scores = [r.get(f"{d}_score") for r in rows if r.get(f"{d}_score") not in ("", None)]
            dim_avgs.append(round(sum(float(s) for s in scores) / len(scores), 1) if scores else "")
        confidence = rows[0].get("data_confidence", "")
        alt = row_num % 2 == 0
        values = [comp, avg_score] + dim_avgs + [confidence, len(rows)]
        for col_i, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_i, value=val)
            cell.border = BORDER
            cell.font = _font(size=9)
            cell.alignment = _align("center")
            cell.fill = _fill(LIGHTER if alt else WHITE)
        # Colour avg score cell
        score_cell = ws.cell(row=excel_row, column=2)
        tier = priority_tier(avg_score)
        if tier == "High":   score_cell.fill = _fill(HIGH_BG); score_cell.font = _font(bold=True, color=HIGH_FG, size=9)
        elif tier == "Medium": score_cell.fill = _fill(MED_BG); score_cell.font = _font(bold=True, color=MED_FG, size=9)
        else:                score_cell.fill = _fill(LOW_BG);  score_cell.font = _font(bold=True, color=LOW_FG, size=9)
        ws.row_dimensions[excel_row].height = 20


def write_initiative_long_list(all_rows: list[dict], path: Path,
                                hypothesis_data: dict | None = None) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    by_competitor: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        by_competitor[row["competitor"]].append(row)

    ws_all = wb.create_sheet(title="All Competitors")
    combined = sorted(all_rows, key=lambda r: (
        r.get("competitor",""), r.get("category",""), -float(r.get("priority_score", 0))
    ))
    _write_long_list_sheet(ws_all, combined, "Initiative Long List — All Competitors")

    for comp_name in sorted(by_competitor.keys()):
        rows = sorted(by_competitor[comp_name],
                      key=lambda r: (r.get("category",""), -float(r.get("priority_score",0))))
        ws = wb.create_sheet(title=comp_name[:28])
        _write_long_list_sheet(ws, rows, f"Initiative Long List — {comp_name}")

    # Comparison Matrix sheet
    ws_matrix = wb.create_sheet(title="Comparison Matrix")
    _write_comparison_matrix_sheet(ws_matrix, all_rows)

    # Hypothesis Tracker sheet
    if hypothesis_data:
        ws_hyp = wb.create_sheet(title="Hypothesis Tracker")
        _write_hypothesis_sheet(ws_hyp, hypothesis_data)

    wb.save(path)
    log.info(f"  Initiative Long List: {path.name}  ({len(all_rows)} services, {len(by_competitor)} competitors)")


def write_detailed_workbook(competitor: str, rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scored Services"
    total_cols = 7 + len(DIMENSIONS) * 2
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    t = ws["A1"]
    t.value = f"AI Services Analysis — {competitor}"
    t.font = _font(bold=True, color=WHITE, size=13)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    base_headers = ["Service","Category","Customer Maturity","AI Classification",
                    "Description","Plain English Summary","Pricing Signals","Client Wins",
                    "Tech Stack","Data Confidence","Priority Score","Priority Tier"]
    dim_headers  = []
    for d in DIMENSIONS:
        dim_headers += [f"{DIM_LABELS[d]} Score", f"{DIM_LABELS[d]} Justification"]
    all_headers = base_headers + dim_headers
    col_widths  = [30,28,24,16,50,50,40,40,40,14,14,12] + [10,38]*len(DIMENSIONS)

    for col_i, (h, w) in enumerate(zip(all_headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font = _font(bold=True, color=WHITE, size=9)
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _align("center", wrap=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    for row_num, row in enumerate(rows, start=1):
        excel_row = row_num + 2
        tier = row.get("priority_tier","")
        alt  = row_num % 2 == 0
        base_values = [row.get("service_name",""), row.get("category",""),
                       row.get("customer_maturity",""), row.get("ai_classification",""),
                       row.get("description",""), row.get("plain_english_summary",""),
                       row.get("pricing_signals",""), row.get("client_wins",""),
                       row.get("tech_stack",""), row.get("data_confidence",""),
                       row.get("priority_score",""), row.get("priority_tier","")]
        dim_values = []
        for d in DIMENSIONS:
            dim_values += [row.get(f"{d}_score",""), row.get(f"{d}_justification","")]
        for col_i, value in enumerate(base_values + dim_values, start=1):
            cell = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border = BORDER
            cell.alignment = _align("left", wrap=True)
            cell.font = _font(size=9)
            cell.fill = _fill(LIGHTER if alt else WHITE)
        for col_i in (11, 12):
            cell = ws.cell(row=excel_row, column=col_i)
            if tier == "High":
                cell.fill = _fill(HIGH_BG); cell.font = _font(bold=True, color=HIGH_FG, size=9)
            elif tier == "Medium":
                cell.fill = _fill(MED_BG);  cell.font = _font(bold=True, color=MED_FG,  size=9)
            else:
                cell.fill = _fill(LOW_BG);  cell.font = _font(bold=True, color=LOW_FG,  size=9)
        ws.row_dimensions[excel_row].height = 40

    wb.save(path)
    log.info(f"  Detailed workbook: {path.name}")

# ═════════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═════════════════════════════════════════════════════════════════════════════

def write_services_md(all_rows: list[dict], path: Path) -> None:
    """Write a plain English markdown summary grouped by competitor."""
    by_competitor: dict[str, list[dict]] = defaultdict(list)
    for row in all_rows:
        by_competitor[row["competitor"]].append(row)

    total = len(all_rows)
    competitors_count = len(by_competitor)
    ts = datetime.now().strftime("%B %Y")

    lines = [
        "# Competitor AI Services — Plain English Guide",
        "",
        "A simple reference for explaining what each competitor offers, without the technical jargon.",
        "",
        "---",
        "",
    ]

    global_num = 1
    for comp in sorted(by_competitor.keys()):
        rows = sorted(by_competitor[comp], key=lambda r: -float(r.get("priority_score", 0)))
        lines.append(f"## {comp}")
        lines.append("")
        for row in rows:
            score = row.get("priority_score", "")
            tier  = row.get("priority_tier", "")
            tier_meaning = {"High": "strong fit — prioritise", "Medium": "moderate fit — consider", "Low": "weak fit — deprioritise"}.get(tier, tier)
            summary = row.get("plain_english_summary", "") or row.get("description", "")
            lines.append(f"**{global_num}. {row.get('service_name', '')}**")
            lines.append(f"*Score: {score}/100 — {tier} priority ({tier_meaning})*")
            lines.append(summary)
            lines.append("")
            global_num += 1
        lines.append("---")
        lines.append("")

    lines.append(f"*Based on scraped website data from {competitors_count} competitor(s) — {ts}.*")

    path.write_text("\n".join(lines), encoding="utf-8")
    log.info(f"  Services MD: {path.name}  ({total} services, {competitors_count} competitors)")


def _write_outputs(master: pd.DataFrame, state: dict) -> None:
    master.to_csv(MASTER_CSV, index=False, encoding="utf-8")
    log.info(f"  Master CSV: {MASTER_CSV.name}  ({len(master)} rows)")
    ts           = datetime.now().strftime("%Y%m%d")
    hyp_data     = load_hypothesis_data()
    ll_path      = OUTPUT_DIR / f"{ts}_initiative_long_list_groq.xlsx"
    write_initiative_long_list(master.to_dict("records"), ll_path, hypothesis_data=hyp_data)
    md_path      = OUTPUT_DIR / f"{ts}_services_summary_groq.md"
    write_services_md(master.to_dict("records"), md_path)
    competitors_done = sum(1 for v in state.values() if not v.get("skipped", False))
    log.info(f"\n{'='*60}")
    log.info(f"  Competitors processed : {competitors_done} / {len(state)} sites tracked")
    log.info(f"  Total services        : {len(master)}")
    for tier in ["High", "Medium", "Low"]:
        n = (master["priority_tier"] == tier).sum()
        log.info(f"  {tier:<8} priority    : {n}")
    log.info(f"  Outputs in            : {OUTPUT_DIR}/")
    log.info(f"{'='*60}\n")


def run(
    competitor_filter: str = "",
    weights_str: str = "",
    max_pages: int = 0,
    ocr_min_len: int = 0,
    dry_run: bool = False,
    rerun_all: bool = False,
) -> pd.DataFrame:
    rerun_all   = rerun_all or os.getenv("RERUN_ALL", "").strip() in ("1", "true", "yes")
    max_pages   = max_pages   or int(os.getenv("MAX_PAGES_PER_SITE", 40))
    ocr_min_len = ocr_min_len or int(os.getenv("OCR_MIN_LENGTH", 30))
    weights     = parse_weights(weights_str)

    OUTPUT_DIR.mkdir(exist_ok=True)

    if rerun_all:
        clear_state()
    state = load_state()

    all_folders = find_site_folders(competitor_filter)
    pending     = [f for f in all_folders if f.name not in state]
    completed   = [f for f in all_folders if f.name in state]

    log.info(f"\nGroq model        : {GROQ_MODEL}")
    log.info(f"Groq API URL      : {GROQ_API_URL}")
    log.info(f"Site folders      : {len(all_folders)}")
    log.info(f"  Already done    : {len(completed)}")
    log.info(f"  Pending         : {len(pending)}")

    if completed:
        log.info("  Completed:")
        for f in completed:
            info = state[f.name]
            log.info(f"    ✓ {competitor_name(f):<25} ({info.get('services','?')} services)")

    if pending:
        log.info("  Pending:")
        for f in pending:
            log.info(f"    … {competitor_name(f)}")

    if dry_run:
        log.info("\nDRY RUN — no model calls made.")
        return pd.DataFrame()

    if not pending:
        log.info("\nNo new folders to process. Rebuilding outputs from existing results...")
        existing = load_existing_results()
        if not existing.empty:
            _write_outputs(existing, state)
        return existing

    new_rows: list[dict] = []

    for folder in pending:
        name = competitor_name(folder)
        log.info(f"\n{'─'*60}")
        log.info(f"  {name}  ({folder.name})")
        log.info(f"{'─'*60}")

        content = load_content(folder, max_pages, ocr_min_len)
        if not content.strip():
            log.warning("  No content — skipping.")
            state[folder.name] = {
                "processed_at": datetime.now().isoformat(timespec="seconds"),
                "services": 0, "competitor": name, "skipped": True,
            }
            save_state(state)
            continue

        log.info(f"  Content: ~{len(content):,} chars")
        confidence = data_confidence(content)
        log.info(f"  Data confidence: {confidence}")
        services = extract_services(name, content)
        if not services:
            log.warning("  No services extracted — skipping.")
            state[folder.name] = {
                "processed_at": datetime.now().isoformat(timespec="seconds"),
                "services": 0, "competitor": name, "skipped": True,
            }
            save_state(state)
            continue

        # Assess hypotheses (one call per competitor, not per service)
        hyp_data = load_hypothesis_data()
        hyp_result = assess_hypotheses(name, content)
        if hyp_result:
            hyp_data[name] = hyp_result
            save_hypothesis_data(hyp_data)

        scored: list[dict] = []
        for i, svc in enumerate(services, 1):
            log.info(f"  Scoring [{i}/{len(services)}]: {svc.get('name','?')}")
            scores = score_service(name, svc)
            scored.append(scores)
            time.sleep(0.3)

        rows = build_rows(name, services, scored, weights, confidence)
        new_rows.extend(rows)

        safe        = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
        detail_path = OUTPUT_DIR / f"{safe}_services_scored_groq.xlsx"
        write_detailed_workbook(name, rows, detail_path)

        state[folder.name] = {
            "processed_at": datetime.now().isoformat(timespec="seconds"),
            "services": len(rows), "competitor": name, "skipped": False,
        }
        save_state(state)
        log.info(f"  ✓ {name} done — {len(rows)} services scored.")

    existing = load_existing_results()
    master   = merge_results(existing, new_rows)
    if master.empty:
        log.warning("No results to write.")
        return master

    _write_outputs(master, state)
    return master


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="AmaliTech Competitor AI Services — Groq Analysis"
    )
    parser.add_argument("--competitor", "-c", default="",
                        help="Filter to a single competitor (e.g. 'andela')")
    parser.add_argument("--weights", "-w", default="",
                        help="Comma-separated weights for 7 dimensions")
    parser.add_argument("--max-pages", "-p", type=int, default=0,
                        help="Max pages per competitor (default: 40)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show state without making model calls")
    parser.add_argument("--rerun-all", action="store_true",
                        help="Ignore state and reprocess everything")
    args = parser.parse_args()

    run(
        competitor_filter=args.competitor,
        weights_str=args.weights,
        max_pages=args.max_pages,
        dry_run=args.dry_run,
        rerun_all=args.rerun_all,
    )


if __name__ == "__main__":
    main()
